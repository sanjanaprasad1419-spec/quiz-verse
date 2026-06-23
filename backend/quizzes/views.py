from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.models import Count, Max, Q
from django.shortcuts import get_object_or_404
from rest_framework import permissions, status, viewsets, parsers
from rest_framework.decorators import action
import csv
from io import StringIO
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, PatternFill

from quizzes.models import Quiz, QuizRegistration, Question, Choice, QuizAttempt, StudentAnswer, FFFAnswer, HotseatAttempt, SwitchCategory, SystemPreferences, BuzzerState, BuzzerPress, Expert
from quizzes.serializers import QuizRegistrationSerializer, QuizSerializer, FFFAnswerSerializer, HotseatAttemptSerializer, QuestionSerializer, EnrolledStudentSerializer, SystemPreferencesSerializer, BuzzerStateSerializer, BuzzerPressSerializer, ExpertSerializer
from quizzes.services import process_mock_payment, register_student_for_quiz
from django.utils import timezone
import random

User = get_user_model()


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == 'admin')


class IsStudentUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == 'student')


def check_admin_quiz_write_access(user, quiz):
    if getattr(user, 'is_super_admin', False):
        return True
    if quiz.created_by == user or quiz.host == user:
        return True
    if getattr(user, 'school', None) and quiz.allowed_schools.filter(id=user.school.id).exists():
        return True
    return False


class AdminQuizViewSet(viewsets.ModelViewSet):
    """
    CRUD for admin to manage quizzes.
    Includes archived quizzes by default.
    """
    permission_classes = [IsAdminUser]
    serializer_class = QuizSerializer
    
    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'is_super_admin', False):
            return Quiz.objects.annotate(
                registered_count=Count('registrations')
            ).all()
        
        queryset = Quiz.objects.annotate(
            registered_count=Count('registrations')
        )

        if getattr(user, 'school', None):
            queryset = queryset.filter(
                Q(created_by=user) | Q(host=user) | Q(allowed_schools=user.school) | Q(allowed_schools__isnull=True)
            )
        else:
            queryset = queryset.filter(
                Q(created_by=user) | Q(host=user) | Q(allowed_schools__isnull=True)
            )
        return queryset.distinct()
        
    def perform_create(self, serializer):
        user = self.request.user
        host_user = serializer.validated_data.get('host', user)
        quiz = serializer.save(created_by=user, host=host_user)
        if not getattr(user, 'is_super_admin', False) and getattr(user, 'school', None):
            quiz.allowed_schools.set([user.school])

    def perform_update(self, serializer):
        user = self.request.user
        quiz = serializer.save()
        if not getattr(user, 'is_super_admin', False) and getattr(user, 'school', None):
            if quiz.allowed_schools.exists():
                quiz.allowed_schools.set([user.school])

    def check_object_permissions(self, request, obj):
        super().check_object_permissions(request, obj)
        if request.method not in permissions.SAFE_METHODS:
            if not getattr(request.user, 'is_super_admin', False):
                if obj.created_by != request.user and obj.host != request.user:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You do not have permission to modify this quiz.")

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        template_type = request.query_params.get('type', 'prelim').strip().lower()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if template_type == 'fff':
            ws.title = "FFF Sequencing Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Option E', 'Option F', 'Option G', 'Option H', 'Option I', 
                'Option J', 'Option K', 'Option L', 'Option M', 'Option N', 'Option O',
                'Correct Sequence (e.g. ECFABGDH)', 'Marks', 
                'Question Type (fff_1/fff_2/fff_3)', 'Category', 'Trivia'
            ]
            sample_row = [
                "Arrange these historical monuments in chronological order of construction (earliest first):",
                "Taj Mahal",
                "Red Fort",
                "Qutub Minar",
                "Gateway of India",
                "Sanchi Stupa",
                "Charminar",
                "Hawa Mahal",
                "Victoria Memorial",
                "", "", "", "", "", "", "",
                "ECFABGDH",
                1,
                "fff_1",
                "History",
                "Sanchi Stupa (E) -> Qutub Minar (C) -> Charminar (F) -> Taj Mahal (A) -> Red Fort (B) -> Hawa Mahal (G) -> Gateway of India (D) -> Victoria Memorial (H). FFF questions strictly require between 4 and 10 options (A to D minimum). Fill options consecutively without gaps. The Correct Sequence must list all defined option letters in order (e.g., ECFABGDH)."
            ]
            filename = "fff_sequencing_template.xlsx"
        elif template_type == 'buzzer':
            ws.title = "Buzzer Round Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Option (A/B/C/D)', 'Marks', 
                'Question Type (buzzer)', 'Category', 'Trivia'
            ]
            sample_row = [
                "Which planet is known as the Red Planet?",
                "Venus",
                "Mars",
                "Jupiter",
                "Saturn",
                "B",
                1,
                "buzzer",
                "Astronomy",
                "Mars is called the Red Planet because of iron oxide on its surface. INSTRUCTIONS: Buzzer Round questions require exactly 4 options (A to D) and the Question Type must be set to 'buzzer'."
            ]
            filename = "buzzer_round_template.xlsx"
        elif template_type == 'hotseat':
            ws.title = "Hotseat MCQ Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Option (A/B/C/D)', 'Marks', 
                'Question Type (hotseat_1/hotseat_2/hotseat_3)', 'Category', 'Trivia'
            ]
            sample_row = [
                "What is the chemical formula of Table Salt?",
                "HCl",
                "H2O",
                "NaCl",
                "CO2",
                "C",
                1,
                "hotseat_1",
                "Science",
                "NaCl stands for Sodium Chloride which is common table salt. INSTRUCTIONS: Hotseat MCQs require exactly 4 options (A to D). Question Type must be 'hotseat_1' (Batch 1), 'hotseat_2' (Batch 2), or 'hotseat_3' (Batch 3) to map questions to the respective hotseat contestant round."
            ]
            filename = "hotseat_quiz_template.xlsx"
        else:
            ws.title = "Preliminary MCQ Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Option (A/B/C/D)', 'Marks', 
                'Question Type (regular)', 'Category', 'Trivia'
            ]
            sample_row = [
                "What is the capital of France?",
                "London",
                "Berlin",
                "Paris",
                "Madrid",
                "C",
                1,
                "regular",
                "General",
                "Paris is the capital of France. INSTRUCTIONS: Preliminary MCQ requires exactly 4 options (A to D) and the Question Type must be set to 'regular'."
            ]
            filename = "preliminary_quiz_template.xlsx"
            
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        ws.append(sample_row)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def upload_questions(self, request, pk=None):
        quiz = self.get_object()
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
            
        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            return Response({"detail": "Please upload an Excel (.xlsx) file."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                return Response({"detail": "No questions found in Excel file."}, status=status.HTTP_400_BAD_REQUEST)
                
            created_count = 0
            error_count = 0
            error_log = []
            seen_questions = set()
            
            max_order = quiz.questions.aggregate(Max('order'))['order__max'] or 0
            
            for idx, row in enumerate(rows[1:], start=2):
                if not row or not any(row):
                    continue
                    
                text = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                if not text:
                    error_log.append({"row": idx, "question": f"Row {idx}", "reason": "Question text is missing."})
                    error_count += 1
                    continue
                    
                if quiz.questions.filter(text=text).exists():
                    error_log.append({"row": idx, "question": text[:40], "reason": "Question text already exists in this quiz."})
                    error_count += 1
                    continue
                    
                if text in seen_questions:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Question text is duplicated inside the Excel file."})
                    error_count += 1
                    continue
                seen_questions.add(text)
                # Detect Excel template format dynamically (Old 10-column vs New 21-column)
                is_old_format = True
                
                # Check if row[18] is a valid question type
                if len(row) > 18 and str(row[18]).strip().lower() in Question.QuestionType.values:
                    is_old_format = False
                elif len(row) > 7 and str(row[7]).strip().lower() in Question.QuestionType.values:
                    is_old_format = True
                else:
                    # Fallback default: if the row has more than 12 columns, assume new format
                    is_old_format = (len(row) < 12)
                
                if is_old_format:
                    # Old 10-column format
                    opt_a = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                    opt_b = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    opt_c = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                    opt_d = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
                    
                    raw_options = [opt_a, opt_b, opt_c, opt_d] + [''] * 11
                    
                    correct_opt = str(row[5]).strip() if len(row) > 5 and row[5] is not None else 'A'
                    marks = int(row[6]) if len(row) > 6 and row[6] is not None else 1
                    q_type = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'regular'
                    category = str(row[8]).strip() if len(row) > 8 and row[8] else 'General'
                    trivia = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ''
                else:
                    # New 21-column format
                    raw_options = []
                    for o_idx in range(15):
                        col_val = str(row[1 + o_idx]).strip() if len(row) > (1 + o_idx) and row[1 + o_idx] is not None else ''
                        raw_options.append(col_val)
                        
                    correct_opt = str(row[16]).strip() if len(row) > 16 and row[16] is not None else 'A'
                    marks = int(row[17]) if len(row) > 17 and row[17] is not None else 1
                    q_type = str(row[18]).strip().lower() if len(row) > 18 and row[18] else 'regular'
                    category = str(row[19]).strip() if len(row) > 19 and row[19] else 'General'
                    trivia = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
                
                # Filter out trailing empty options to find the non-empty option set
                non_empty_options = [o for o in raw_options if o]
                num_options = len(non_empty_options)
                
                # Ensure options are contiguous without gaps (only check up to num_options)
                has_gap = False
                for idx_opt in range(num_options):
                    if not raw_options[idx_opt]:
                        has_gap = True
                        break
                if has_gap:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Options must be contiguous starting from Option A without empty slots."})
                    error_count += 1
                    continue
                
                if q_type not in Question.QuestionType.values:
                    q_type = 'regular'
                
                # Check if the round for this question type is enabled
                if q_type == 'regular' and not quiz.has_prelim_round:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Preliminary MCQ round is disabled for this quiz."})
                    error_count += 1
                    continue
                elif q_type.startswith('fff_') and not quiz.has_fff_round:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Fastest Finger First round is disabled for this quiz."})
                    error_count += 1
                    continue
                elif q_type == 'buzzer' and not quiz.has_buzzer_round:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Buzzer round is disabled for this quiz."})
                    error_count += 1
                    continue
                elif (q_type.startswith('hotseat_') or q_type == 'switch') and not quiz.has_hotseat_round:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Hotseat round is disabled for this quiz."})
                    error_count += 1
                    continue
                
                # Validation checks specific to FFF vs regular MCQ
                if q_type.startswith('fff_'):
                    if num_options < 4 or num_options > 10:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Fastest Finger First questions must have between 4 and 10 options. Got {num_options}."})
                        error_count += 1
                        continue
                    
                    clean_seq = [ch.upper() for ch in correct_opt if ch.isalpha()]
                    expected_letters = [chr(ord('A') + i) for i in range(num_options)]
                    
                    if len(clean_seq) != num_options or set(clean_seq) != set(expected_letters):
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Correct sequence for FFF must specify exactly the non-empty option letters A to {chr(ord('A') + num_options - 1)} in order (case-insensitive). Got '{correct_opt}'."})
                        error_count += 1
                        continue
                    
                    seq_map = {letter: rank for rank, letter in enumerate(clean_seq, 1)}
                else:
                    # MCQ strictly requires exactly 4 options
                    if num_options != 4:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Standard MCQ / Preliminary / Hotseat questions must have exactly 4 options. Got {num_options}."})
                        error_count += 1
                        continue
                    
                    correct_opt_upper = correct_opt.upper()
                    expected_letters = ['A', 'B', 'C', 'D']
                    if correct_opt_upper not in expected_letters:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Correct option letter '{correct_opt}' is invalid. Must be A, B, C, or D (case-insensitive)."})
                        error_count += 1
                        continue
                
                with transaction.atomic():
                    question = Question.objects.create(
                        quiz=quiz, text=text, order=max_order + created_count + 1, marks=marks,
                        question_type=q_type, category=category, trivia=trivia
                    )
                    
                    if q_type.startswith('fff_'):
                        for letter_idx, letter in enumerate([chr(ord('A') + i) for i in range(num_options)]):
                            Choice.objects.create(
                                question=question,
                                text=raw_options[letter_idx],
                                is_correct=False,
                                correct_order=seq_map[letter]
                            )
                    else:
                        for letter_idx, letter in enumerate(expected_letters):
                            Choice.objects.create(
                                question=question,
                                text=raw_options[letter_idx],
                                is_correct=(correct_opt_upper == letter),
                                correct_order=None
                            )
                    
                created_count += 1
                
            return Response({
                "detail": f"Successfully imported {created_count} questions.",
                "success_count": created_count,
                "error_count": error_count,
                "errors": error_log
            })
            
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        quiz = self.get_object()
        questions = quiz.questions.all().prefetch_related('choices')
        
        data = []
        for q in questions:
            choices_data = []
            for c in q.choices.all():
                choices_data.append({
                    "id": c.id,
                    "text": c.text,
                    "is_correct": c.is_correct,
                    "correct_order": c.correct_order
                })
            data.append({
                "id": q.id,
                "text": q.text,
                "order": q.order,
                "marks": q.marks,
                "question_type": q.question_type,
                "category": q.category,
                "trivia": q.trivia,
                "choices": choices_data
            })
        return Response(data)

    @action(detail=True, methods=['get'])
    def switch_categories(self, request, pk=None):
        quiz = self.get_object()
        categories = quiz.switch_categories.all().select_related('question')
        
        data = []
        for c in categories:
            choices_data = []
            q_data = None
            if c.question:
                for choice in c.question.choices.all():
                    choices_data.append({
                        "id": choice.id,
                        "text": choice.text,
                        "is_correct": choice.is_correct
                    })
                q_data = {
                    "id": c.question.id,
                    "text": c.question.text,
                    "choices": choices_data
                }
                
            img_url = c.image.url if c.image else None
            if img_url and request:
                img_url = request.build_absolute_uri(img_url)
                
            data.append({
                "id": c.id,
                "name": c.name,
                "image": img_url,
                "question": q_data
            })
        return Response(data)

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser, parsers.FormParser])
    def save_switch_category(self, request, pk=None):
        quiz = self.get_object()
        
        category_id = request.data.get('category_id')
        if not category_id and quiz.switch_categories.count() >= 6:
            return Response({"detail": "You can configure a maximum of 6 switch categories."}, status=400)
            
        name = request.data.get('name', '').strip()
        if not name:
            return Response({"detail": "Category name is required."}, status=400)
            
        question_text = request.data.get('question_text', '').strip()
        if not question_text:
            return Response({"detail": "Question text is required."}, status=400)
            
        choice_a = request.data.get('choice_a', '').strip()
        choice_b = request.data.get('choice_b', '').strip()
        choice_c = request.data.get('choice_c', '').strip()
        choice_d = request.data.get('choice_d', '').strip()
        correct_choice = request.data.get('correct_choice', '').strip().upper()
        
        if not all([choice_a, choice_b, choice_c, choice_d]) or correct_choice not in ['A', 'B', 'C', 'D']:
            return Response({"detail": "All 4 options and a correct selection (A/B/C/D) are required."}, status=400)
            
        with transaction.atomic():
            if category_id:
                category = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
                category.name = name
                if 'image' in request.FILES:
                    category.image = request.FILES['image']
                category.save()
            else:
                image_file = request.FILES.get('image')
                category = SwitchCategory.objects.create(
                    quiz=quiz,
                    name=name,
                    image=image_file
                )
                
            if category.question:
                question = category.question
                question.text = question_text
                question.category = name
                question.save()
                question.choices.all().delete()
            else:
                question = Question.objects.create(
                    quiz=quiz,
                    text=question_text,
                    question_type=Question.QuestionType.SWITCH,
                    category=name,
                    order=0
                )
                category.question = question
                category.save()
                
            Choice.objects.create(question=question, text=choice_a, is_correct=(correct_choice == 'A'))
            Choice.objects.create(question=question, text=choice_b, is_correct=(correct_choice == 'B'))
            Choice.objects.create(question=question, text=choice_c, is_correct=(correct_choice == 'C'))
            Choice.objects.create(question=question, text=choice_d, is_correct=(correct_choice == 'D'))
            
        return Response({"detail": "Switch category and question saved successfully."})

    @action(detail=True, methods=['post'])
    def delete_switch_category(self, request, pk=None):
        quiz = self.get_object()
        category_id = request.data.get('category_id')
        if not category_id:
            return Response({"detail": "Category ID is required."}, status=400)
            
        category = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        with transaction.atomic():
            if category.question:
                category.question.delete()
            category.delete()
            
        return Response({"detail": "Switch category deleted successfully."})

    @action(detail=True, methods=['get'])
    def get_experts(self, request, pk=None):
        quiz = self.get_object()
        experts = quiz.experts.all()
        data = []
        for e in experts:
            photo_url = e.photo.url if e.photo else None
            if photo_url and request:
                photo_url = request.build_absolute_uri(photo_url)
            data.append({
                "id": e.id,
                "name": e.name,
                "designation": e.designation,
                "photo": photo_url
            })
        return Response(data)

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser, parsers.FormParser])
    def save_expert(self, request, pk=None):
        quiz = self.get_object()
        expert_id = request.data.get('expert_id')
        name = request.data.get('name', '').strip()
        designation = request.data.get('designation', '').strip()
        if not name:
            return Response({"detail": "Expert name is required."}, status=400)
            
        if expert_id:
            expert = get_object_or_404(Expert, quiz=quiz, id=expert_id)
            expert.name = name
            expert.designation = designation
            if 'photo' in request.FILES:
                expert.photo = request.FILES['photo']
            expert.save()
        else:
            if quiz.experts.count() >= 5:
                return Response({"detail": "Maximum of 5 experts allowed for a quiz."}, status=400)
            photo_file = request.FILES.get('photo')
            expert = Expert.objects.create(
                quiz=quiz,
                name=name,
                designation=designation,
                photo=photo_file
            )
        
        photo_url = expert.photo.url if expert.photo else None
        if photo_url and request:
            photo_url = request.build_absolute_uri(photo_url)
            
        return Response({
            "detail": "Expert saved successfully.",
            "expert": {
                "id": expert.id,
                "name": expert.name,
                "designation": expert.designation,
                "photo": photo_url
            }
        })

    @action(detail=True, methods=['post'])
    def delete_expert(self, request, pk=None):
        quiz = self.get_object()
        expert_id = request.data.get('expert_id')
        if not expert_id:
            return Response({"detail": "Expert ID is required."}, status=400)
        expert = get_object_or_404(Expert, quiz=quiz, id=expert_id)
        expert.delete()
        return Response({"detail": "Expert deleted successfully."})

    @action(detail=True, methods=['post'])
    def add_question(self, request, pk=None):
        quiz = self.get_object()
        text = request.data.get('text', '').strip()
        if not text:
            return Response({"detail": "Question text is required."}, status=400)
            
        q_type = request.data.get('question_type', 'regular')
        
        # Validation checks based on enabled quiz rounds
        if q_type == 'regular' and not quiz.has_prelim_round:
            return Response({"detail": "Preliminary MCQ round is disabled for this quiz."}, status=400)
        elif q_type.startswith('fff_') and not quiz.has_fff_round:
            return Response({"detail": "Fastest Finger First round is disabled for this quiz."}, status=400)
        elif q_type == 'buzzer' and not quiz.has_buzzer_round:
            return Response({"detail": "Buzzer round is disabled for this quiz."}, status=400)
        elif (q_type.startswith('hotseat_') or q_type == 'switch') and not quiz.has_hotseat_round:
            return Response({"detail": "Hotseat round is disabled for this quiz."}, status=400)
        category = request.data.get('category', 'General')
        marks = int(request.data.get('marks', 1))
        trivia = request.data.get('trivia', '')
        choices_data = request.data.get('choices', [])
        
        max_order = quiz.questions.aggregate(Max('order'))['order__max'] or 0
        
        with transaction.atomic():
            question = Question.objects.create(
                quiz=quiz, text=text, order=max_order + 1,
                marks=marks, question_type=q_type, category=category, trivia=trivia
            )
            for c in choices_data:
                Choice.objects.create(
                    question=question,
                    text=c.get('text', '').strip(),
                    is_correct=c.get('is_correct', False),
                    correct_order=c.get('correct_order')
                )
        return Response({"detail": "Question added successfully.", "id": question.id})

    @action(detail=False, methods=['post'])
    def edit_question(self, request):
        question_id = request.data.get('id')
        question = get_object_or_404(Question, id=question_id)
        
        text = request.data.get('text', '').strip()
        if not text:
            return Response({"detail": "Question text is required."}, status=400)
            
        q_type = request.data.get('question_type', 'regular')
        
        # Validation checks based on enabled quiz rounds
        quiz = question.quiz
        if q_type == 'regular' and not quiz.has_prelim_round:
            return Response({"detail": "Preliminary MCQ round is disabled for this quiz."}, status=400)
        elif q_type.startswith('fff_') and not quiz.has_fff_round:
            return Response({"detail": "Fastest Finger First round is disabled for this quiz."}, status=400)
        elif q_type == 'buzzer' and not quiz.has_buzzer_round:
            return Response({"detail": "Buzzer round is disabled for this quiz."}, status=400)
        elif (q_type.startswith('hotseat_') or q_type == 'switch') and not quiz.has_hotseat_round:
            return Response({"detail": "Hotseat round is disabled for this quiz."}, status=400)
        category = request.data.get('category', 'General')
        marks = int(request.data.get('marks', 1))
        trivia = request.data.get('trivia', '')
        choices_data = request.data.get('choices', [])
        
        with transaction.atomic():
            question.text = text
            question.question_type = q_type
            question.category = category
            question.marks = marks
            question.trivia = trivia
            question.save()
            
            question.choices.all().delete()
            for c in choices_data:
                Choice.objects.create(
                    question=question,
                    text=c.get('text', '').strip(),
                    is_correct=c.get('is_correct', False),
                    correct_order=c.get('correct_order')
                )
        return Response({"detail": "Question updated successfully."})

    @action(detail=False, methods=['post'])
    def delete_question(self, request):
        question_id = request.data.get('id')
        question = get_object_or_404(Question, id=question_id)
        question.delete()
        return Response({"detail": "Question deleted successfully."})

    @action(detail=True, methods=['post'])
    def update_stage(self, request, pk=None):
        quiz = self.get_object()
        stage = request.data.get('stage')
        if not stage or stage not in Quiz.Stage.values:
            return Response({"detail": "Invalid stage provided."}, status=400)
        quiz.current_stage = stage
        quiz.save(update_fields=['current_stage'])
        return Response(QuizSerializer(quiz).data)

    @action(detail=True, methods=['get'])
    def host_hotseat_question(self, request, pk=None):
        """Admin host view: see current question with correct answers, trivia, and contestant's preselection."""
        quiz = self.get_object()
        stage = quiz.current_stage
        
        hotseat_player = None
        batch_num = None
        q_type = None
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"active": False, "detail": "No hotseat stage active."})
        
        if not hotseat_player:
            return Response({"active": False, "detail": "No hotseat player promoted."})
        
        attempt = HotseatAttempt.objects.filter(quiz=quiz, student=hotseat_player, batch_number=batch_num).first()
        if not attempt:
            return Response({"active": False, "detail": "No hotseat attempt found."})
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            # Return question data even after round ends for answer key display
            questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
            if attempt.current_question_index < len(questions):
                question = questions[attempt.current_question_index]
                choices = list(question.choices.all())
                question_data = {
                    "id": question.id,
                    "text": question.text,
                    "category": question.category,
                    "trivia": question.trivia,
                    "choices": [{"id": c.id, "text": c.text, "is_correct": c.is_correct} for c in choices]
                }
            else:
                question_data = None
            return Response({
                "active": False,
                "completed": True,
                "status": attempt.status,
                "score": attempt.score,
                "contestant_name": hotseat_player.full_name,
                "question": question_data
            })
        
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"active": False, "completed": True, "status": "completed", "score": attempt.score})
        
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        return Response({
            "active": True,
            "current_index": attempt.current_question_index,
            "total_questions": len(questions),
            "score": attempt.score,
            "contestant_name": hotseat_player.full_name,
            "preselected_choice_id": attempt.preselected_choice_id,
            "lifelines": {
                "5050_used": attempt.lifeline_5050_used,
                "poll_used": attempt.lifeline_poll_used,
                "switch_used": attempt.lifeline_switch_used,
                "expert_used": attempt.lifeline_expert_used
            },
            "lifeline_request_status": attempt.lifeline_request_status,
            "pending_lifeline_type": attempt.pending_lifeline_type,
            "pending_lifeline_switch_category": attempt.pending_lifeline_switch_category,
            "approved_lifeline_data": attempt.approved_lifeline_data,
            "timer_is_paused": attempt.timer_is_paused,
            "options_visible": attempt.options_visible,
            "showing_question": attempt.showing_question,
            "question": {
                "id": question.id,
                "text": question.text,
                "category": question.category,
                "trivia": question.trivia,
                "choices": [{"id": c.id, "text": c.text, "is_correct": c.is_correct} for c in choices]
            }
        })

    @action(detail=True, methods=['post'])
    def host_lock_answer(self, request, pk=None):
        """Admin host action: lock the contestant's preselected answer and process scoring."""
        quiz = self.get_object()
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "No hotseat stage active."}, status=400)
        
        if not hotseat_player:
            return Response({"detail": "No hotseat player promoted."}, status=400)
        
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
        
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "All questions completed."}, status=400)
        
        question = questions[attempt.current_question_index]
        choice_id = attempt.preselected_choice_id
        if not choice_id:
            return Response({"detail": "Contestant has not selected any option yet."}, status=400)
        
        selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
        is_correct = selected_choice.is_correct if selected_choice else False
        correct_choice = Choice.objects.filter(question=question, is_correct=True).first()
        
        if is_correct:
            current_points = SCORE_LADDER[attempt.current_question_index] if attempt.current_question_index < len(SCORE_LADDER) else 0
            attempt.score = current_points
            attempt.current_question_index += 1
            attempt.preselected_choice = None
            attempt.current_question_switched = False
            
            # Reset option visibility and hold next question until explicitly pushed by Host
            attempt.showing_question = False
            attempt.options_visible = False
            attempt.timer_is_paused = False
            
            completed = attempt.current_question_index >= len(questions)
            if completed:
                attempt.status = HotseatAttempt.Status.COMPLETED
                attempt.completed_at = timezone.now()
                self._host_save_score(quiz, batch_num, attempt.score, "completed")
            else:
                self._host_save_score(quiz, batch_num, attempt.score, "playing")
            
            attempt.save()
            return Response({
                "correct": True,
                "correct_choice_id": correct_choice.id if correct_choice else None,
                "selected_choice_id": choice_id,
                "current_points": attempt.score,
                "next_index": attempt.current_question_index,
                "completed": completed,
                "trivia": question.trivia,
                "message": f"All {len(questions)} questions completed! Final score: {attempt.score} pts" if completed else f"Correct! {hotseat_player.full_name} has reached Question {attempt.current_question_index + 1}!"
            })
        else:
            checkpoint_score = 0
            fail_index = attempt.current_question_index
            if fail_index >= 10:
                checkpoint_score = 320000
            elif fail_index >= 5:
                checkpoint_score = 10000
            
            attempt.score = checkpoint_score
            attempt.status = HotseatAttempt.Status.FAILED
            attempt.completed_at = timezone.now()
            attempt.preselected_choice = None
            attempt.save()
            
            self._host_save_score(quiz, batch_num, attempt.score, "failed")
            
            return Response({
                "correct": False,
                "correct_choice_id": correct_choice.id if correct_choice else None,
                "selected_choice_id": choice_id,
                "checkpoint_points": checkpoint_score,
                "completed": True,
                "trivia": question.trivia,
                "message": f"Incorrect! The correct answer was '{correct_choice.text if correct_choice else 'N/A'}'. Score drops to checkpoint: {checkpoint_score} pts"
            })

    def _host_save_score(self, quiz, batch_num, score, status_str):
        if batch_num == 1:
            quiz.hotseat_score_1 = score
            quiz.hotseat_status_1 = status_str
            quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
        elif batch_num == 2:
            quiz.hotseat_score_2 = score
            quiz.hotseat_status_2 = status_str
            quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
        elif batch_num == 3:
            quiz.hotseat_score_3 = score
            quiz.hotseat_status_3 = status_str
            quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])

    @action(detail=True, methods=['post'])
    def set_batches(self, request, pk=None):
        quiz = self.get_object()
        batch_1 = request.data.get('batch_1', [])
        batch_2 = request.data.get('batch_2', [])
        batch_3 = request.data.get('batch_3', [])
        
        if not batch_1 or not batch_2 or not batch_3:
            attempts = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).order_by('-score', 'completed_at')
            student_ids = [att.student_id for att in attempts]
            total_participants = len(student_ids)
            
            if total_participants == 1:
                batch_1 = [student_ids[0]]
                batch_2, batch_3 = [], []
                top_selected = list(batch_1)
            elif total_participants == 2:
                batch_1 = [student_ids[0]]
                batch_2 = [student_ids[1]]
                batch_3 = []
                top_selected = list(batch_1) + list(batch_2)
            else:
                top_30_percent_count = int(round(total_participants * 0.30))
                batch_size = top_30_percent_count // 3
                if batch_size < 1:
                    batch_size = 1
                total_to_select = batch_size * 3
                top_selected = student_ids[:total_to_select]
                batch_1 = top_selected[0:batch_size]
                batch_2 = top_selected[batch_size:batch_size*2]
                batch_3 = top_selected[batch_size*2:total_to_select]
            
            quiz.top_30_selected = top_selected
        else:
            quiz.top_30_selected = list(batch_1) + list(batch_2) + list(batch_3)

        quiz.batch_1_players = batch_1
        quiz.batch_2_players = batch_2
        quiz.batch_3_players = batch_3
        quiz.save(update_fields=['top_30_selected', 'batch_1_players', 'batch_2_players', 'batch_3_players'])
        return Response(QuizSerializer(quiz).data)


    @action(detail=True, methods=['get'])
    def fff_results(self, request, pk=None):
        quiz = self.get_object()
        stage = quiz.current_stage
        if stage == Quiz.Stage.FFF_BATCH_1 or stage == Quiz.Stage.HOTSEAT_BATCH_1:
            batch_num = 1
            q_type = Question.QuestionType.FFF_1
        elif stage == Quiz.Stage.FFF_BATCH_2 or stage == Quiz.Stage.HOTSEAT_BATCH_2:
            batch_num = 2
            q_type = Question.QuestionType.FFF_2
        elif stage == Quiz.Stage.FFF_BATCH_3 or stage == Quiz.Stage.HOTSEAT_BATCH_3:
            batch_num = 3
            q_type = Question.QuestionType.FFF_3
        else:
            return Response({"detail": "FFF is not active or completed for this quiz stage."}, status=400)
            
        fff_question = Question.objects.filter(quiz=quiz, question_type=q_type).first()
        if not fff_question:
            return Response({"detail": "FFF question not found for this batch."}, status=404)
            
        answers = FFFAnswer.objects.filter(question=fff_question, batch_number=batch_num).select_related('student')
        
        serialized = FFFAnswerSerializer(answers, many=True).data
        
        for ans_data, ans_obj in zip(serialized, answers):
            ans_data['is_correct'] = ans_obj.is_correct or (ans_obj.selected_choice.is_correct if ans_obj.selected_choice else False)
            
        serialized.sort(key=lambda x: (not x['is_correct'], x['time_taken_seconds']))
        
        return Response({
            "question": QuestionSerializer(fff_question).data,
            "results": serialized
        })

    @action(detail=True, methods=['post'])
    def promote_hotseat(self, request, pk=None):
        quiz = self.get_object()
        stage = quiz.current_stage
        student_id = request.data.get('student_id')
        if not student_id:
            return Response({"detail": "Student ID is required."}, status=400)
            
        student = get_object_or_404(User, id=student_id)
        
        if stage == Quiz.Stage.FFF_BATCH_1 or stage == Quiz.Stage.HOTSEAT_BATCH_1:
            quiz.hotseat_player_1 = student
            quiz.hotseat_status_1 = "playing"
            quiz.save(update_fields=['hotseat_player_1', 'hotseat_status_1'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=1)
        elif stage == Quiz.Stage.FFF_BATCH_2 or stage == Quiz.Stage.HOTSEAT_BATCH_2:
            quiz.hotseat_player_2 = student
            quiz.hotseat_status_2 = "playing"
            quiz.save(update_fields=['hotseat_player_2', 'hotseat_status_2'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=2)
        elif stage == Quiz.Stage.FFF_BATCH_3 or stage == Quiz.Stage.HOTSEAT_BATCH_3:
            quiz.hotseat_player_3 = student
            quiz.hotseat_status_3 = "playing"
            quiz.save(update_fields=['hotseat_player_3', 'hotseat_status_3'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=3)
        else:
            return Response({"detail": "Promotion is not allowed in this stage."}, status=400)
            
        return Response(QuizSerializer(quiz).data)

    @action(detail=True, methods=['get'])
    def prelim_scores(self, request, pk=None):
        quiz = self.get_object()
        # Fetch all attempts (both completed and in-progress), ordered by:
        # 1. Completed first, then in-progress
        # 2. Higher score first
        # 3. Earlier completion time first (for completed attempts)
        from django.db.models import Case, When, Value, IntegerField
        attempts = QuizAttempt.objects.filter(quiz=quiz).select_related('student').annotate(
            completion_order=Case(
                When(completed_at__isnull=False, then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('completion_order', '-score', 'completed_at')
        
        total_regular_questions = quiz.questions.filter(question_type=Question.QuestionType.REGULAR).count()
        
        data = []
        for idx, att in enumerate(attempts, 1):
            reg = QuizRegistration.objects.filter(quiz=quiz, student=att.student).first()
            answers = att.answers.select_related('selected_choice').all()
            correct_count = sum(1 for ans in answers if ans.selected_choice and ans.selected_choice.is_correct)
            incorrect_count = sum(1 for ans in answers if not ans.selected_choice or not ans.selected_choice.is_correct)
            is_completed = att.completed_at is not None
            data.append({
                "rank": idx,
                "student_id": att.student.id,
                "student_name": att.student.full_name,
                "player_id": reg.player_id if reg else "",
                "score": att.score,
                "time_taken": (att.completed_at - att.started_at).total_seconds() if att.completed_at and att.started_at else None,
                "correct_count": correct_count,
                "incorrect_count": incorrect_count,
                "completed": is_completed,
                "questions_answered": att.current_question_index,
                "total_questions": total_regular_questions
            })
        return Response(data)

    @action(detail=True, methods=['get'])
    def enrolled_students(self, request, pk=None):
        quiz = self.get_object()
        registrations = quiz.registrations.select_related('student').all()
        serializer = EnrolledStudentSerializer(registrations, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def remove_registration(self, request, pk=None):
        quiz = self.get_object()
        registration_id = request.data.get('registration_id')
        if not registration_id:
            return Response({"detail": "registration_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        reg = QuizRegistration.objects.filter(id=registration_id, quiz=quiz).first()
        if not reg:
            return Response({"detail": "Registration not found."}, status=status.HTTP_404_NOT_FOUND)
        
        student = reg.student
        
        # Delete any associated quiz attempt and answers so the student can start fresh
        QuizAttempt.objects.filter(quiz=quiz, student=student).delete()
        
        # Delete FFF answers if any
        FFFAnswer.objects.filter(quiz=quiz, student=student).delete()
        
        # Delete Hotseat attempts if any
        HotseatAttempt.objects.filter(quiz=quiz, student=student).delete()
        
        # Clear hotseat player references on the quiz if this student was assigned
        update_fields = []
        if quiz.hotseat_player_1 == student:
            quiz.hotseat_player_1 = None
            quiz.hotseat_status_1 = ""
            update_fields.extend(['hotseat_player_1', 'hotseat_status_1'])
        if quiz.hotseat_player_2 == student:
            quiz.hotseat_player_2 = None
            quiz.hotseat_status_2 = ""
            update_fields.extend(['hotseat_player_2', 'hotseat_status_2'])
        if quiz.hotseat_player_3 == student:
            quiz.hotseat_player_3 = None
            quiz.hotseat_status_3 = ""
            update_fields.extend(['hotseat_player_3', 'hotseat_status_3'])
        
        # Remove student from batch player lists
        if student.id in quiz.batch_1_players:
            quiz.batch_1_players.remove(student.id)
            if 'batch_1_players' not in update_fields:
                update_fields.append('batch_1_players')
        if student.id in quiz.batch_2_players:
            quiz.batch_2_players.remove(student.id)
            if 'batch_2_players' not in update_fields:
                update_fields.append('batch_2_players')
        if student.id in quiz.batch_3_players:
            quiz.batch_3_players.remove(student.id)
            if 'batch_3_players' not in update_fields:
                update_fields.append('batch_3_players')
        
        if update_fields:
            quiz.save(update_fields=update_fields)
        
        # Delete the registration itself
        reg.delete()
        
        return Response({
            "detail": f"Registration for {student.full_name} has been fully removed. All quiz data (attempts, answers, hotseat, FFF) has been cleared. They can now re-register.",
            "removed_student_name": student.full_name,
            "removed_student_email": student.email
        })

    @action(detail=True, methods=['post'])
    def enroll_student_manual(self, request, pk=None):
        quiz = self.get_object()
        email = request.data.get('email', '').strip().lower()
        full_name = request.data.get('full_name', '').strip()
        college_id = request.data.get('roll_number', '').strip() or request.data.get('college_id', '').strip()
        payment_status = request.data.get('payment_status', 'paid').strip().lower()

        if not email or not full_name or not college_id:
            return Response({"detail": "Email, full name, and roll number are required."}, status=status.HTTP_400_BAD_REQUEST)

        prefs = SystemPreferences.get_solo()
        if prefs.auto_approve_registrations:
            payment_status = QuizRegistration.PaymentStatus.PAID
        elif payment_status not in [QuizRegistration.PaymentStatus.PAID, QuizRegistration.PaymentStatus.PENDING]:
            payment_status = QuizRegistration.PaymentStatus.PAID

        college_id = college_id.upper().strip()
        email = email.strip().lower()

        # Query existing student accounts strictly by roll_number or email case-insensitively
        from django.db.models import Q
        user = User.objects.filter(
            Q(roll_number__iexact=college_id) | 
            Q(college_id__iexact=college_id) | 
            Q(email__iexact=email)
        ).first()

        if not user:
            from users.models import School, Program, Branch, StudentProfile
            user = User.objects.create(
                email=email,
                full_name=full_name,
                college_id=college_id.upper(),
                roll_number=college_id.upper(),
                role=User.Role.STUDENT
            )
            user.set_password("KBC123")
            user.save()

            # Ensure they have a student profile created automatically
            school = School.objects.first()
            if not school:
                school = School.objects.create(school_name="Default School", school_code="DEFAULT_SCH")
            
            program = Program.objects.filter(school=school).first()
            if not program:
                program = Program.objects.create(school=school, program_name="Default Program", program_code="DEFAULT_PROG")
            
            branch = Branch.objects.filter(program=program).first()
            if not branch:
                branch = Branch.objects.create(program=program, branch_name="Default Branch", branch_code="DEFAULT_BR")

            StudentProfile.objects.create(
                user=user,
                school=school,
                program=program,
                branch=branch,
                year="1"
            )

        if user.role != User.Role.STUDENT:
            return Response(
                {"detail": "The specified user is not a student account."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Make sure the user's roll_number and college_id are capitalized to prevent cased duplicate lookups
                changed = False
                if user.roll_number != user.roll_number.upper():
                    user.roll_number = user.roll_number.upper()
                    changed = True
                if user.college_id != user.college_id.upper():
                    user.college_id = user.college_id.upper()
                    changed = True
                if changed:
                    user.save(update_fields=['roll_number', 'college_id', 'updated_at'])

                if QuizRegistration.objects.filter(student=user, quiz=quiz).exists():
                    reg = QuizRegistration.objects.get(student=user, quiz=quiz)
                    reg.payment_status = payment_status
                    reg.save()
                    return Response({
                        "detail": "Student is already registered for this quiz. Registration status updated.",
                        "registration": EnrolledStudentSerializer(reg).data
                    })

                max_seq = QuizRegistration.objects.filter(
                    quiz=quiz,
                    sequence_number__isnull=False
                ).aggregate(Max('sequence_number'))['sequence_number__max']
                next_seq = 1 if max_seq is None else max_seq + 1

                reg = QuizRegistration.objects.create(
                    student=user,
                    quiz=quiz,
                    payment_status=payment_status,
                    sequence_number=next_seq,
                    player_id=f"PLAYER {next_seq:03d}"
                )

            return Response({
                "detail": f"Successfully enrolled student {full_name}.",
                "registration": EnrolledStudentSerializer(reg).data
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def bulk_enroll_students(self, request, pk=None):
        quiz = self.get_object()
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
            
        file = request.FILES['file']
        filename = file.name.lower()
        
        is_xlsx = filename.endswith('.xlsx')
        is_csv = filename.endswith('.csv')
        
        if not is_xlsx and not is_csv:
            return Response({"detail": "Please upload an Excel (.xlsx) or CSV (.csv) file."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            rows = []
            if is_xlsx:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            elif is_csv:
                file_data = file.read().decode('utf-8')
                csv_data = csv.reader(StringIO(file_data))
                rows = list(csv_data)
 
            if len(rows) <= 1:
                return Response({"detail": "No student records found in the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)
                
            enrolled_count = 0
            skipped_count = 0
            
            from users.models import School, Program, Branch, StudentProfile
            prefs = SystemPreferences.get_solo()
            
            with transaction.atomic():
                for idx, row in enumerate(rows[1:], start=1):
                    if not row or not any(row):
                        continue
                        
                    full_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                    email = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ''
                    college_id = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    pay_status = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else 'paid'
                    
                    if not email or not full_name or not college_id:
                        skipped_count += 1
                        continue
                        
                    if prefs.auto_approve_registrations:
                        pay_status = QuizRegistration.PaymentStatus.PAID
                    elif pay_status not in [QuizRegistration.PaymentStatus.PAID, QuizRegistration.PaymentStatus.PENDING]:
                        pay_status = QuizRegistration.PaymentStatus.PAID
 
                    user, created = User.objects.get_or_create(
                        email=email,
                        defaults={
                            'full_name': full_name,
                            'college_id': college_id,
                            'roll_number': college_id,
                            'role': User.Role.STUDENT
                        }
                    )
                    if created:
                        user.set_password("KBC123")
                        user.save()
                    else:
                        changed = False
                        if not user.college_id:
                            user.college_id = college_id
                            changed = True
                        if not user.roll_number:
                            user.roll_number = college_id
                            changed = True
                        if changed:
                            user.save()
                            
                    if not hasattr(user, 'student_profile'):
                        school = School.objects.first()
                        if not school:
                            school = School.objects.create(school_name="Default School", school_code="DEFAULT_SCH")
                        
                        program = Program.objects.filter(school=school).first()
                        if not program:
                            program = Program.objects.create(school=school, program_name="Default Program", program_code="DEFAULT_PROG")
                        
                        branch = Branch.objects.filter(program=program).first()
                        if not branch:
                            branch = Branch.objects.create(program=program, branch_name="Default Branch", branch_code="DEFAULT_BR")
 
                        StudentProfile.objects.create(
                            user=user,
                            school=school,
                            program=program,
                            branch=branch,
                            year=StudentProfile.Year.FIRST
                        )
                        
                    if QuizRegistration.objects.filter(student=user, quiz=quiz).exists():
                        reg = QuizRegistration.objects.get(student=user, quiz=quiz)
                        reg.payment_status = pay_status
                        reg.save()
                        continue
                        
                    max_seq = QuizRegistration.objects.filter(
                        quiz=quiz,
                        sequence_number__isnull=False
                    ).aggregate(Max('sequence_number'))['sequence_number__max']
                    next_seq = 1 if max_seq is None else max_seq + 1
 
                    QuizRegistration.objects.create(
                        student=user,
                        quiz=quiz,
                        payment_status=pay_status,
                        sequence_number=next_seq,
                        player_id=f"PLAYER {next_seq:03d}"
                    )
                    enrolled_count += 1
                    
            return Response({
                "detail": f"Successfully enrolled {enrolled_count} students. Skipped {skipped_count} invalid records."
            })
            
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
    @action(detail=False, methods=['get'])
    def download_enrollment_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Enrollment Template"
        
        headers = ['Full Name', 'Email', 'Roll Number', 'Payment Status (paid/pending)']
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        sample_row = [
            "John Doe",
            "johndoe@quizverse.edu",
            "ROLL001",
            "paid"
        ]
        ws.append(sample_row)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_enrollment_template.xlsx"'
        wb.save(response)
        return response


class MyQuizRegistrationView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        reg = QuizRegistration.objects.filter(quiz=quiz, student=request.user).first()
        if not reg:
            return Response({"registered": False})
        return Response({
            "registered": True,
            "id": reg.id,
            "player_id": reg.player_id,
            "payment_status": reg.payment_status,
            "event_password_required": True
        })


class QuizAttemptStartView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if quiz.status != Quiz.Status.LIVE:
            return Response({"detail": "The quiz is not live yet."}, status=403)
        reg = get_object_or_404(QuizRegistration, quiz=quiz, student=request.user)
        if reg.payment_status != 'paid':
            return Response({"detail": "Payment required."}, status=403)
            
        attempt, created = QuizAttempt.objects.get_or_create(student=request.user, quiz=quiz)
        if created or not attempt.question_ids:
            import random
            q_ids = list(quiz.questions.filter(question_type=Question.QuestionType.REGULAR).values_list('id', flat=True))
            random.shuffle(q_ids)
            attempt.question_ids = q_ids
            attempt.save(update_fields=['question_ids'])
        return Response({"attempt_id": attempt.id, "current_index": attempt.current_question_index})


class QuizAttemptNextQuestionView(APIView):
    permission_classes = [IsStudentUser]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if quiz.status != Quiz.Status.LIVE:
            return Response({"detail": "The quiz is not live yet."}, status=403)
        attempt = get_object_or_404(QuizAttempt, student=request.user, quiz=quiz)
        
        if attempt.question_ids:
            preserved_order = {id_val: i for i, id_val in enumerate(attempt.question_ids)}
            questions_queryset = quiz.questions.filter(id__in=attempt.question_ids)
            questions = sorted(list(questions_queryset), key=lambda q: preserved_order.get(q.id, 99999))
        else:
            questions = list(quiz.questions.filter(question_type=Question.QuestionType.REGULAR).order_by('order', 'id'))
            
        if attempt.current_question_index >= len(questions):
            return Response({"completed": True})
            
        question = questions[attempt.current_question_index]
        choices = [{"id": c.id, "text": c.text} for c in question.choices.all()]
        
        return Response({
            "completed": False,
            "question": {
                "id": question.id,
                "text": question.text,
                "marks": question.marks,
                "order": question.order,
                "choices": choices,
                "total_questions": len(questions),
                "current_index": attempt.current_question_index
            }
        })


class QuizAttemptSubmitAnswerView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if quiz.status != Quiz.Status.LIVE:
            return Response({"detail": "The quiz is not live yet."}, status=403)
        attempt = get_object_or_404(QuizAttempt, student=request.user, quiz=quiz)
        
        if attempt.question_ids:
            preserved_order = {id_val: i for i, id_val in enumerate(attempt.question_ids)}
            questions_queryset = quiz.questions.filter(id__in=attempt.question_ids)
            questions = sorted(list(questions_queryset), key=lambda q: preserved_order.get(q.id, 99999))
        else:
            questions = list(quiz.questions.filter(question_type=Question.QuestionType.REGULAR).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "Quiz already completed."}, status=400)
            
        question = questions[attempt.current_question_index]
        choice_id = request.data.get('choice_id')
        time_taken = request.data.get('time_taken', 0)
        
        selected_choice = None
        if choice_id:
            selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
            if selected_choice and selected_choice.is_correct:
                attempt.score += question.marks
                
        StudentAnswer.objects.create(
            attempt=attempt,
            question=question,
            selected_choice=selected_choice,
            time_taken_seconds=time_taken
        )
        
        attempt.current_question_index += 1
        attempt.save()
        
        is_completed = attempt.current_question_index >= len(questions)
        if is_completed:
            from django.utils import timezone
            attempt.completed_at = timezone.now()
            attempt.save()
            
        correct_choice = Choice.objects.filter(question=question, is_correct=True).first()
        correct_choice_data = None
        if correct_choice:
            correct_choice_data = {
                "id": correct_choice.id,
                "text": correct_choice.text
            }
            
        return Response({
            "correct": selected_choice.is_correct if selected_choice else False,
            "completed": is_completed,
            "correct_choice": correct_choice_data,
            "trivia": question.trivia
        })


class AdminStatsView(APIView):
    """
    Returns platform-wide or school-specific statistics for the admin dashboard.
    """
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        admin_user = request.user
        
        # Base querysets
        students = User.objects.filter(role='student')
        quizzes = Quiz.objects.filter(is_archived=False)
        registrations = QuizRegistration.objects.all()
        
        # Enforce school boundaries for school admins
        if not admin_user.is_super_admin:
            if admin_user.school:
                students = students.filter(student_profile__school=admin_user.school)
                quizzes = quizzes.filter(allowed_schools=admin_user.school)
                registrations = registrations.filter(student__student_profile__school=admin_user.school)
            else:
                students = students.none()
                quizzes = quizzes.none()
                registrations = registrations.none()
                
        total_students = students.count()
        total_quizzes = quizzes.count()
        active_quizzes = quizzes.filter(visible_to_students=True).count()
        total_registrations = registrations.count()
        
        # Year-wise breakdowns directly from database
        from django.db.models import Count
        from users.models import StudentProfile
        
        profile_qs = StudentProfile.objects.all()
        if not admin_user.is_super_admin and admin_user.school:
            profile_qs = profile_qs.filter(school=admin_user.school)
            
        year_counts = profile_qs.values('year').annotate(count=Count('id'))
        year_breakdown = {
            "1": 0,
            "2": 0,
            "3": 0,
            "4": 0
        }
        for yc in year_counts:
            if yc['year'] in year_breakdown:
                year_breakdown[yc['year']] = yc['count']
                
        return Response({
            "total_students": total_students,
            "total_quizzes": total_quizzes,
            "active_quizzes": active_quizzes,
            "total_registrations": total_registrations,
            "year1_count": year_breakdown["1"],
            "year2_count": year_breakdown["2"],
            "year3_count": year_breakdown["3"],
            "year4_count": year_breakdown["4"],
        })


class PublishedQuizListView(APIView):
    """
    Lists all published, non-archived quizzes for students.
    """
    permission_classes = [IsStudentUser]
    
    def get(self, request):
        quizzes = Quiz.objects.annotate(
            registered_count=Count('registrations')
        ).filter(
            visible_to_students=True,
            is_archived=False
        )
        
        profile = getattr(request.user, 'student_profile', None)
        if profile and profile.school_id:
            from django.db.models import Q
            quizzes = quizzes.filter(Q(allowed_schools__id=profile.school_id) | Q(allowed_schools__isnull=True))
            quizzes = quizzes.exclude(~Q(created_by__school_id=profile.school_id) & Q(created_by__is_super_admin=False) & Q(created_by__school__isnull=False))
            
        return Response(QuizSerializer(quizzes, many=True, context={'request': request}).data)



class QuizDetailView(APIView):
    """
    Detailed view of a single quiz, available to authenticated users.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(
            Quiz.objects.annotate(registered_count=Count('registrations')), 
            pk=pk
        )
        
        # Prevent students from viewing hidden/archived quizzes or quizzes they are not eligible for
        if request.user.role == 'student':
            if not quiz.visible_to_students or quiz.is_archived:
                return Response({"detail": "Quiz not found or not available."}, status=status.HTTP_404_NOT_FOUND)
            
            profile = getattr(request.user, 'student_profile', None)
            if profile and quiz.allowed_schools.exists() and not quiz.allowed_schools.filter(id=profile.school_id).exists():
                return Response({"detail": "Quiz not found or not available."}, status=status.HTTP_404_NOT_FOUND)
            
        return Response(QuizSerializer(quiz, context={'request': request}).data)


class StudentRegistrationView(APIView):
    """
    Handles student registration for a quiz.
    """
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        try:
            registration = register_student_for_quiz(request.user, quiz)
            return Response(
                QuizRegistrationSerializer(registration).data, 
                status=status.HTTP_201_CREATED
            )
        except DjangoValidationError as e:
            raise ValidationError({"detail": str(e.message) if hasattr(e, 'message') else str(e)})


class MockPaymentView(APIView):
    """
    Handles the simulated payment success for a pending registration.
    """
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        registration = get_object_or_404(
            QuizRegistration, 
            quiz_id=pk, 
            student=request.user
        )
        
        try:
            processed_registration = process_mock_payment(registration)
            return Response(QuizRegistrationSerializer(processed_registration).data)
        except DjangoValidationError as e:
            raise ValidationError({"detail": str(e.message) if hasattr(e, 'message') else str(e)})


class MyRegistrationsView(APIView):
    """
    Returns the authenticated student's registrations.
    """
    permission_classes = [IsStudentUser]
    
    def get(self, request):
        registrations = QuizRegistration.objects.select_related('quiz').filter(
            student=request.user
        )
        return Response(QuizRegistrationSerializer(registrations, many=True).data)

from quizzes.models import Team
from quizzes.serializers import TeamSerializer
from django.db.models import Q

class StudentTeamViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TeamSerializer
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            if user.is_super_admin:
                return Team.objects.all()
            if user.school:
                return Team.objects.filter(quiz__allowed_schools=user.school).distinct()
            return Team.objects.none()
        else:
            # Teams for quizzes the user is registered for
            return Team.objects.filter(
                Q(leader=user) | Q(members=user) | Q(quiz__registrations__student=user)
            ).distinct()
        
    def validate_team_members(self, quiz, leader, emails, current_team_id=None):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        from quizzes.models import QuizRegistration, Team
        from rest_framework.exceptions import ValidationError
        
        # 1. Check for blank emails
        for idx, email in enumerate(emails):
            if not email or not str(email).strip():
                raise ValidationError({"detail": f"Member {idx + 1} email is required."})
        
        # Check that leader's email matches the first email
        if emails[0].lower().strip() != leader.email.lower().strip():
            raise ValidationError({"detail": "The creator/leader of the team must be Member 1."})

        unique_emails = set(e.lower().strip() for e in emails)
        if len(unique_emails) < 4:
            raise ValidationError({"detail": "All 4 members in the team must have unique emails."})

        users = []
        for idx, email in enumerate(emails):
            email_clean = email.lower().strip()
            # Check user existence
            try:
                user = User.objects.get(email__iexact=email_clean)
            except User.DoesNotExist:
                raise ValidationError({"detail": f"Member {idx + 1} email ({email}) is not registered in the system."})
            
            # Check registration for this quiz
            if not QuizRegistration.objects.filter(quiz=quiz, student=user).exists():
                raise ValidationError({"detail": f"Member {idx + 1} ({user.full_name}) is not registered for this quiz."})
                
            # Check if already in another team for this quiz (leader or member)
            other_teams = Team.objects.filter(quiz=quiz).filter(Q(leader=user) | Q(members=user))
            if current_team_id:
                other_teams = other_teams.exclude(pk=current_team_id)
            if other_teams.exists():
                raise ValidationError({"detail": f"Member {idx + 1} ({user.full_name}) is already registered in another team."})
                
            users.append(user)
            
        return users

    def perform_create(self, serializer):
        if self.request.user.role != 'student':
            raise ValidationError({"detail": "Only student users can register teams."})
            
        quiz_id = self.request.data.get('quiz')
        quiz = get_object_or_404(Quiz, pk=quiz_id)
        if not QuizRegistration.objects.filter(quiz=quiz, student=self.request.user).exists():
            raise ValidationError({"detail": "You must be registered for this quiz to create a team."})
            
        emails = [
            self.request.data.get('member1_email'),
            self.request.data.get('member2_email'),
            self.request.data.get('member3_email'),
            self.request.data.get('member4_email')
        ]
        
        users = self.validate_team_members(quiz, self.request.user, emails)
        instance = serializer.save(
            leader=self.request.user, 
            quiz=quiz,
            member1_name=users[0].full_name,
            member1_email=users[0].email,
            member2_name=users[1].full_name,
            member2_email=users[1].email,
            member3_name=users[2].full_name,
            member3_email=users[2].email,
            member4_name=users[3].full_name,
            member4_email=users[3].email,
        )
        instance.members.set([users[1], users[2], users[3]])
            
    def perform_update(self, serializer):
        from rest_framework.exceptions import PermissionDenied
        if self.request.user.role != 'student':
            raise PermissionDenied("Only student users can edit teams.")
            
        team = self.get_object()
        if team.leader != self.request.user:
            raise PermissionDenied("Only the team leader can edit the team.")
            
        emails = [
            self.request.data.get('member1_email') or team.member1_email,
            self.request.data.get('member2_email') or team.member2_email,
            self.request.data.get('member3_email') or team.member3_email,
            self.request.data.get('member4_email') or team.member4_email
        ]
        
        users = self.validate_team_members(team.quiz, team.leader, emails, current_team_id=team.id)
        instance = serializer.save(
            member1_name=users[0].full_name,
            member1_email=users[0].email,
            member2_name=users[1].full_name,
            member2_email=users[1].email,
            member3_name=users[2].full_name,
            member3_email=users[2].email,
            member4_name=users[3].full_name,
            member4_email=users[3].email,
        )
        instance.members.set([users[1], users[2], users[3]])
            
    def perform_destroy(self, instance):
        from rest_framework.exceptions import PermissionDenied
        if self.request.user.role != 'student':
            raise PermissionDenied("Only student users can delete teams.")
            
        if instance.leader != self.request.user:
            raise PermissionDenied("Only the team leader can delete the team.")
        instance.delete()
        
    @action(detail=True, methods=['post'])
    def join(self, request, pk=None):
        if request.user.role != 'student':
            return Response({"detail": "Only student users can join teams."}, status=400)
            
        team = self.get_object()
        if not QuizRegistration.objects.filter(quiz=team.quiz, student=request.user).exists():
            return Response({"detail": "You must be registered for this quiz to join this team."}, status=400)
            
        from quizzes.models import Team as TeamModel
        other_teams = TeamModel.objects.filter(quiz=team.quiz).filter(Q(leader=request.user) | Q(members=request.user))
        if other_teams.exists():
            return Response({"detail": "You are already in a team for this quiz."}, status=400)
            
        if team.members.count() + 1 >= 4:
            return Response({"detail": "This team has reached its maximum size of 4 members."}, status=400)
            
        team.members.add(request.user)
        return Response({"detail": "Successfully joined team!"})


class StudentRegisteredPlayersView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        # Check if requesting user is registered
        if not QuizRegistration.objects.filter(quiz=quiz, student=request.user).exists():
            return Response({"detail": "You must be registered for this quiz to view other participants."}, status=403)
            
        # Get all registered students
        registrations = QuizRegistration.objects.filter(quiz=quiz).select_related('student')
        
        data = []
        for r in registrations:
            data.append({
                "id": r.student.id,
                "full_name": r.student.full_name,
                "email": r.student.email,
            })
        return Response(data)




# Student KBC Event views
class VerifyQuizAccessView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        player_id = request.data.get('player_id', '').strip()
        password = request.data.get('event_password', '').strip()
        
        reg = QuizRegistration.objects.filter(quiz=quiz, student=request.user).first()
        if not reg or reg.payment_status != 'paid':
            return Response({"detail": "You are not registered or paid for this quiz."}, status=403)
            
        if not reg.arena_password:
            reg.save()
            
        if reg.player_id.lower() != player_id.lower():
            return Response({"detail": "Invalid Player ID."}, status=400)
            
        if not password:
            return Response({"detail": "Password is required to enter the arena."}, status=400)
            
        if reg.arena_password != password:
            return Response({"detail": "Invalid Event Password. Please enter the unique password shown on your dashboard."}, status=400)
            
        return Response({"success": True, "detail": "Access granted to the arena."})


class QuizLiveStateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        # Auto-populate batches if empty and stage is batch_selection or later FFF/Hotseat stages
        if quiz.current_stage != Quiz.Stage.REGULAR and not quiz.batch_1_players and not quiz.batch_2_players and not quiz.batch_3_players:
            attempts = list(QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).order_by('-score', 'completed_at'))
            if attempts:
                student_ids = [att.student_id for att in attempts]
                total_participants = len(student_ids)
                
                if total_participants == 1:
                    batch_1 = [student_ids[0]]
                    batch_2, batch_3 = [], []
                    top_selected = list(batch_1)
                elif total_participants == 2:
                    batch_1 = [student_ids[0]]
                    batch_2 = [student_ids[1]]
                    batch_3 = []
                    top_selected = list(batch_1) + list(batch_2)
                else:
                    top_30_percent_count = int(round(total_participants * 0.30))
                    batch_size = top_30_percent_count // 3
                    if batch_size < 1:
                        batch_size = 1
                    total_to_select = batch_size * 3
                    top_selected = student_ids[:total_to_select]
                    batch_1 = top_selected[0:batch_size]
                    batch_2 = top_selected[batch_size:batch_size*2]
                    batch_3 = top_selected[batch_size*2:total_to_select]
                
                quiz.batch_1_players = batch_1
                quiz.batch_2_players = batch_2
                quiz.batch_3_players = batch_3
                quiz.top_30_selected = top_selected
                quiz.save(update_fields=['top_30_selected', 'batch_1_players', 'batch_2_players', 'batch_3_players'])

        user = request.user
        
        role = "spectator"
        batch_number = None
        is_in_active_batch = False
        hotseat_attempt_data = None
        fff_question_data = None
        fff_answered = False
        
        if user.role == 'student':
            user_id = user.id
            is_testing_quiz = "test" in quiz.title.lower()
            
            if user_id in quiz.batch_1_players:
                role = "batch_player"
                batch_number = 1
            elif user_id in quiz.batch_2_players:
                role = "batch_player"
                batch_number = 2
            elif user_id in quiz.batch_3_players:
                role = "batch_player"
                batch_number = 3
            elif is_testing_quiz:
                role = "batch_player"
                stage = quiz.current_stage
                if stage in [Quiz.Stage.FFF_BATCH_2, Quiz.Stage.HOTSEAT_BATCH_2]:
                    batch_number = 2
                elif stage in [Quiz.Stage.FFF_BATCH_3, Quiz.Stage.HOTSEAT_BATCH_3]:
                    batch_number = 3
                else:
                    batch_number = 1
                
            stage = quiz.current_stage
            if (stage == Quiz.Stage.FFF_BATCH_1 and batch_number == 1) or \
               (stage == Quiz.Stage.FFF_BATCH_2 and batch_number == 2) or \
               (stage == Quiz.Stage.FFF_BATCH_3 and batch_number == 3):
                is_in_active_batch = True
                
            if (stage == Quiz.Stage.HOTSEAT_BATCH_1 and quiz.hotseat_player_1 == user) or \
               (stage == Quiz.Stage.HOTSEAT_BATCH_2 and quiz.hotseat_player_2 == user) or \
               (stage == Quiz.Stage.HOTSEAT_BATCH_3 and quiz.hotseat_player_3 == user):
                role = "hotseat_player"
                
            # Load FFF question if active
            fff_question_data = None
            fff_answered = False
            fff_q_type = None
            fff_batch = None
            if stage == Quiz.Stage.FFF_BATCH_1:
                fff_q_type = Question.QuestionType.FFF_1
                fff_batch = 1
            elif stage == Quiz.Stage.FFF_BATCH_2:
                fff_q_type = Question.QuestionType.FFF_2
                fff_batch = 2
            elif stage == Quiz.Stage.FFF_BATCH_3:
                fff_q_type = Question.QuestionType.FFF_3
                fff_batch = 3
                
            if fff_q_type:
                fff_q = Question.objects.filter(quiz=quiz, question_type=fff_q_type).first()
                if fff_q:
                    fff_question_data = {
                        "id": fff_q.id,
                        "text": fff_q.text,
                        "choices": [{"id": c.id, "text": c.text} for c in fff_q.choices.all()]
                    }
                    fff_answered = FFFAnswer.objects.filter(
                        quiz=quiz, student=user, batch_number=fff_batch, question=fff_q
                    ).exists()

        # Load active hotseat attempt data for all roles (students, hosts, and spectators)
        stage = quiz.current_stage
        active_hotseat_player = None
        active_batch = None
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            active_hotseat_player = quiz.hotseat_player_1
            active_batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            active_hotseat_player = quiz.hotseat_player_2
            active_batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            active_hotseat_player = quiz.hotseat_player_3
            active_batch = 3
            
        if active_hotseat_player:
            attempt = HotseatAttempt.objects.filter(quiz=quiz, student=active_hotseat_player, batch_number=active_batch).first()
            if attempt:
                # Real-time Spectator Audience Poll Aggregation
                if attempt.pending_lifeline_type == 'poll' and attempt.lifeline_request_status == 'approved':
                    from quizzes.models import SpectatorPollVote
                    from django.utils.dateparse import parse_datetime
                    
                    approved_data = attempt.approved_lifeline_data or {}
                    poll_start_time_str = approved_data.get('poll_start_time')
                    votes_closed = approved_data.get('votes_closed', False)
                    
                    if poll_start_time_str and not votes_closed:
                        poll_start_time = parse_datetime(poll_start_time_str)
                        if poll_start_time:
                            elapsed = (timezone.now() - poll_start_time).total_seconds()
                            
                            # Determine question type
                            if active_batch == 1:
                                q_type = Question.QuestionType.HOTSEAT_1
                            elif active_batch == 2:
                                q_type = Question.QuestionType.HOTSEAT_2
                            else:
                                q_type = Question.QuestionType.HOTSEAT_3
                                
                            questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
                            
                            if attempt.current_question_index < len(questions):
                                question = questions[attempt.current_question_index]
                                choices = list(question.choices.all())
                                
                                if elapsed < 15.0:
                                    # Active window: compute live votes
                                    votes_count = {}
                                    total_votes = 0
                                    for c in choices:
                                        cnt = SpectatorPollVote.objects.filter(attempt=attempt, question=question, choice=c).count()
                                        votes_count[c.id] = cnt
                                        total_votes += cnt
                                    
                                    live_votes = {}
                                    if total_votes > 0:
                                        for c_id, count in votes_count.items():
                                            live_votes[c_id] = round((count / total_votes) * 100)
                                    else:
                                        for c in choices:
                                            live_votes[c.id] = 0
                                            
                                    approved_data['votes'] = live_votes
                                    attempt.approved_lifeline_data = approved_data
                                    attempt.save(update_fields=['approved_lifeline_data'])
                                else:
                                    # Timer expired: finalize votes with KBC fallback
                                    votes_count = {}
                                    total_votes = 0
                                    for c in choices:
                                        cnt = SpectatorPollVote.objects.filter(attempt=attempt, question=question, choice=c).count()
                                        votes_count[c.id] = cnt
                                        total_votes += cnt
                                        
                                    final_votes = {}
                                    if total_votes > 0:
                                        for c_id, count in votes_count.items():
                                            final_votes[c_id] = round((count / total_votes) * 100)
                                    else:
                                        # Fallback to KBC randomized correct-answer-weighted mock generator
                                        correct_choice = next(c for c in choices if c.is_correct)
                                        import random
                                        correct_votes = random.randint(55, 75)
                                        remaining_votes = 100 - correct_votes
                                        
                                        incorrect_choices = [c for c in choices if not c.is_correct]
                                        random.shuffle(incorrect_choices)
                                        
                                        final_votes[correct_choice.id] = correct_votes
                                        
                                        if len(incorrect_choices) >= 3:
                                            v1 = random.randint(5, max(5, remaining_votes - 10))
                                            remaining_votes -= v1
                                            v2 = random.randint(2, max(2, remaining_votes - 5))
                                            remaining_votes -= v2
                                            v3 = remaining_votes
                                            
                                            final_votes[incorrect_choices[0].id] = v1
                                            final_votes[incorrect_choices[1].id] = v2
                                            final_votes[incorrect_choices[2].id] = v3
                                        else:
                                            for idx, inc in enumerate(incorrect_choices):
                                                if idx == len(incorrect_choices) - 1:
                                                    final_votes[inc.id] = remaining_votes
                                                else:
                                                    v = random.randint(5, max(5, remaining_votes // 2))
                                                    final_votes[inc.id] = v
                                                    remaining_votes -= v
                                                    
                                    approved_data['votes'] = final_votes
                                    approved_data['votes_closed'] = True
                                    attempt.approved_lifeline_data = approved_data
                                    attempt.save(update_fields=['approved_lifeline_data'])
                                    
                hotseat_attempt_data = HotseatAttemptSerializer(attempt).data
                    
        live_participants = quiz.registrations.count()
        total_questions = quiz.questions.filter(question_type=Question.QuestionType.REGULAR).count()
        overall_total_questions = quiz.questions.count()
                    
        # Resolve batch player names in a single database query
        all_ids = set((quiz.batch_1_players or []) + (quiz.batch_2_players or []) + (quiz.batch_3_players or []))
        user_map = {}
        if all_ids:
            users = User.objects.filter(id__in=all_ids)
            user_map = {u.id: u.full_name for u in users}

        def resolve_players_list(id_list):
            if not id_list:
                return []
            return [{"id": pid, "name": user_map.get(pid, f"Player ID: {pid}")} for pid in id_list]

        b1_resolved = resolve_players_list(quiz.batch_1_players)
        b2_resolved = resolve_players_list(quiz.batch_2_players)
        b3_resolved = resolve_players_list(quiz.batch_3_players)

        # Buzzer state resolution
        buzzer_state_data = None
        if quiz.current_stage == Quiz.Stage.BUZZER_ROUND:
            b_state, created = BuzzerState.objects.get_or_create(quiz=quiz)
            updated = False
            if created or not b_state.buzzer_mappings:
                b_state.buzzer_mappings = {str(i): {"name": f"Podium {i}", "score": 0} for i in range(1, 16)}
                updated = True
            if not b_state.current_question:
                q = Question.objects.filter(quiz=quiz, question_type=Question.QuestionType.BUZZER).order_by('order', 'id').first()
                if q:
                    b_state.current_question = q
                    updated = True
            if b_state.is_timer_running and b_state.timer_started_at:
                elapsed = (timezone.now() - b_state.timer_started_at).total_seconds()
                if elapsed >= b_state.answer_timer_limit:
                    b_state.is_timer_running = False
                    b_state.buzzers_locked = True
                    updated = True
            if updated:
                b_state.save()
            buzzer_state_data = BuzzerStateSerializer(b_state).data

        prefs = SystemPreferences.get_solo()

        return Response({
            "quiz_id": quiz.id,
            "title": quiz.title,
            "status": quiz.status,
            "intro_title": quiz.intro_title or "Kaun Banega Crorepati",
            "current_stage": quiz.current_stage,
            "student_role": role,
            "batch_number": batch_number,
            "is_in_active_batch": is_in_active_batch,
            "hotseat_attempt": hotseat_attempt_data,
            "buzzer_state": buzzer_state_data,
            "stage_display": quiz.get_current_stage_display(),
            "fff_question": fff_question_data,
            "fff_answered": fff_answered,
            "live_participants": live_participants,
            "total_questions": total_questions,
            "overall_total_questions": overall_total_questions,
            "batch_1_players": b1_resolved,
            "batch_2_players": b2_resolved,
            "batch_3_players": b3_resolved,
            "prelim_mcq_timer": prefs.prelim_mcq_timer,
            "fff_speed_timer": prefs.fff_speed_timer,
            "hotseat_q1_q5_limit": prefs.hotseat_q1_q5_limit,
            "hotseat_q6_q10_limit": prefs.hotseat_q6_q10_limit,
            "auto_approve_registrations": prefs.auto_approve_registrations,
            "expert_timer_limit": getattr(prefs, 'expert_timer_limit', 30),
        })


class FFFSubmitView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.FFF_BATCH_1:
            batch_num = 1
            q_type = Question.QuestionType.FFF_1
            players = quiz.batch_1_players
        elif stage == Quiz.Stage.FFF_BATCH_2:
            batch_num = 2
            q_type = Question.QuestionType.FFF_2
            players = quiz.batch_2_players
        elif stage == Quiz.Stage.FFF_BATCH_3:
            batch_num = 3
            q_type = Question.QuestionType.FFF_3
            players = quiz.batch_3_players
        else:
            return Response({"detail": "Fastest Finger First is not active at this stage."}, status=400)
            
        is_testing_quiz = "test" in quiz.title.lower()
        if request.user.id not in players and not is_testing_quiz:
            return Response({"detail": "You are not in the active batch for this Fastest Finger First round."}, status=403)
            
        fff_question = Question.objects.filter(quiz=quiz, question_type=q_type).first()
        if not fff_question:
            return Response({"detail": "FFF question not found."}, status=404)
            
        if FFFAnswer.objects.filter(quiz=quiz, student=request.user, batch_number=batch_num, question=fff_question).exists():
            return Response({"detail": "You have already submitted your answer for this round."}, status=400)
            
        selected_sequence = request.data.get('selected_sequence', [])
        time_taken = float(request.data.get('time_taken', 0.0))
        
        # If selected_sequence is not sent but choice_id is, wrap it in a list
        choice_id = request.data.get('choice_id')
        if not selected_sequence and choice_id:
            selected_sequence = [choice_id]
            
        correct_choices = Choice.objects.filter(question=fff_question, correct_order__isnull=False).order_by('correct_order')
        
        if correct_choices.exists():
            correct_sequence = [c.id for c in correct_choices]
            try:
                student_seq = [int(x) for x in selected_sequence]
            except (ValueError, TypeError):
                student_seq = []
            is_sequence_correct = (student_seq == correct_sequence)
        else:
            # Fallback to old single-choice logic
            first_choice_id = selected_sequence[0] if selected_sequence else None
            selected_choice = None
            if first_choice_id:
                selected_choice = Choice.objects.filter(question=fff_question, id=first_choice_id).first()
            is_sequence_correct = selected_choice.is_correct if selected_choice else False
        
        first_choice = None
        if selected_sequence:
            try:
                first_choice_id = int(selected_sequence[0])
                first_choice = Choice.objects.filter(question=fff_question, id=first_choice_id).first()
            except (ValueError, TypeError, IndexError):
                pass
        
        answer = FFFAnswer.objects.create(
            quiz=quiz,
            student=request.user,
            batch_number=batch_num,
            question=fff_question,
            selected_choice=first_choice,
            time_taken_seconds=time_taken,
            is_correct=is_sequence_correct,
            submitted_sequence=",".join(map(str, selected_sequence))
        )
        
        return Response({
            "submitted": True,
            "correct": is_sequence_correct
        })


class HotseatQuestionView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No hotseat contestant has been promoted yet."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"completed": True, "status": attempt.status, "score": attempt.score})
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if not questions:
            return Response({"detail": "No hotseat questions have been uploaded for this batch."}, status=404)
            
        if attempt.current_question_index >= len(questions):
            attempt.status = HotseatAttempt.Status.COMPLETED
            attempt.completed_at = timezone.now()
            attempt.save()
            return Response({"completed": True, "status": attempt.status, "score": attempt.score})
            
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        return Response({
            "completed": False,
            "current_index": attempt.current_question_index,
            "total_questions": len(questions),
            "score": attempt.score,
            "preselected_choice_id": attempt.preselected_choice_id,
            "question": {
                "id": question.id,
                "text": question.text,
                "category": question.category,
                "order": question.order,
                "choices": [{"id": c.id, "text": c.text} for c in choices]
            }
        })


class HotseatPreselectView(APIView):
    """Allows the hotseat player to save their selected choice without locking it."""
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active."}, status=400)
        
        if request.user != hotseat_player:
            return Response({"detail": "You are not the hotseat contestant."}, status=403)
        
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt is not active."}, status=400)
        
        choice_id = request.data.get('choice_id')
        if choice_id:
            choice = Choice.objects.filter(id=choice_id).first()
            attempt.preselected_choice = choice
        else:
            attempt.preselected_choice = None
        attempt.save(update_fields=['preselected_choice'])
        
        return Response({"detail": "Selection updated.", "preselected_choice_id": choice_id})


SCORE_LADDER = [
    1000,     # Q1
    2000,     # Q2
    3000,     # Q3
    5000,     # Q4
    10000,    # Q5 (Checkpoint 1)
    20000,    # Q6
    40000,    # Q7
    80000,    # Q8
    160000,   # Q9
    320000,   # Q10 (Checkpoint 2)
    640000,   # Q11
    1250000,  # Q12
    2500000,  # Q13
    5000000,  # Q14
    10000000  # Q15
]

class HotseatSubmitView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "All questions completed."}, status=400)
            
        question = questions[attempt.current_question_index]
        choice_id = request.data.get('choice_id')
        
        selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
        is_correct = selected_choice.is_correct if selected_choice else False
        
        if is_correct:
            current_points = SCORE_LADDER[attempt.current_question_index]
            attempt.score = current_points
            attempt.current_question_index += 1
            attempt.current_question_switched = False
            
            if attempt.current_question_index >= len(questions):
                attempt.status = HotseatAttempt.Status.COMPLETED
                attempt.completed_at = timezone.now()
                self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "completed")
            else:
                self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "playing")
                
            attempt.save()
            return Response({
                "correct": True,
                "current_points": attempt.score,
                "next_index": attempt.current_question_index,
                "completed": attempt.status == HotseatAttempt.Status.COMPLETED
            })
        else:
            checkpoint_score = 0
            fail_index = attempt.current_question_index
            if fail_index >= 10:
                checkpoint_score = 320000  # Drop to Checkpoint 2 score (Q10)
            elif fail_index >= 5:
                checkpoint_score = 10000   # Drop to Checkpoint 1 score (Q5)
            else:
                checkpoint_score = 0
                
            attempt.score = checkpoint_score
            attempt.status = HotseatAttempt.Status.FAILED
            attempt.completed_at = timezone.now()
            attempt.save()
            
            self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "failed")
            
            return Response({
                "correct": False,
                "correct_choice_id": Choice.objects.filter(question=question, is_correct=True).values_list('id', flat=True).first(),
                "checkpoint_points": attempt.score,
                "completed": True
            })
            
    def save_quiz_hotseat_score(self, quiz, batch_num, score, status):
        if batch_num == 1:
            quiz.hotseat_score_1 = score
            quiz.hotseat_status_1 = status
            quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
        elif batch_num == 2:
            quiz.hotseat_score_2 = score
            quiz.hotseat_status_2 = status
            quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
        elif batch_num == 3:
            quiz.hotseat_score_3 = score
            quiz.hotseat_status_3 = status
            quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])


class HotseatLifelineView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if not attempt.options_visible:
            return Response({"detail": "Lifelines can only be requested after choices are shown by the host."}, status=400)
            
        if attempt.current_question_index >= 14:
            return Response({"detail": "Lifelines are no longer available on the 15th question."}, status=400)
            
        lifeline = request.data.get('lifeline') or request.data.get('lifeline_type')
        if not lifeline or lifeline not in ['5050', 'poll', 'switch', 'expert']:
            return Response({"detail": "Invalid lifeline provided."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        if lifeline == '5050':
            if attempt.lifeline_5050_used:
                return Response({"detail": "50:50 lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            incorrect_choices = [c for c in choices if not c.is_correct]
            
            random.shuffle(incorrect_choices)
            eliminated = incorrect_choices[:2]
            
            attempt.lifeline_5050_used = True
            attempt.save()
            
            return Response({
                "lifeline": "5050",
                "eliminated_choice_ids": [c.id for c in eliminated]
            })
            
        elif lifeline == 'poll':
            if attempt.lifeline_poll_used:
                return Response({"detail": "Audience Poll lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            
            correct_votes = random.randint(55, 75)
            remaining_votes = 100 - correct_votes
            
            incorrect_choices = [c for c in choices if not c.is_correct]
            random.shuffle(incorrect_choices)
            
            poll_results = {}
            poll_results[correct_choice.id] = correct_votes
            
            if len(incorrect_choices) >= 3:
                v1 = random.randint(5, max(5, remaining_votes - 10))
                remaining_votes -= v1
                v2 = random.randint(2, max(2, remaining_votes - 5))
                remaining_votes -= v2
                v3 = remaining_votes
                
                poll_results[incorrect_choices[0].id] = v1
                poll_results[incorrect_choices[1].id] = v2
                poll_results[incorrect_choices[2].id] = v3
            else:
                for idx, inc in enumerate(incorrect_choices):
                    if idx == len(incorrect_choices) - 1:
                        poll_results[inc.id] = remaining_votes
                    else:
                        v = random.randint(5, max(5, remaining_votes // 2))
                        poll_results[inc.id] = v
                        remaining_votes -= v
                        
            attempt.lifeline_poll_used = True
            attempt.save()
            
            return Response({
                "lifeline": "poll",
                "votes": poll_results
            })
            
        elif lifeline == 'switch':
            if attempt.lifeline_switch_used:
                return Response({"detail": "Switch Question lifeline already used."}, status=400)
                
            category = request.data.get('category', 'General')
            
            replacement = Question.objects.filter(
                quiz=quiz,
                question_type=q_type,
                category__iexact=category
            ).exclude(id=question.id).first()
            
            if not replacement:
                replacement = Question.objects.filter(
                    quiz=quiz,
                    question_type=q_type
                ).exclude(id=question.id).first()
                
            if not replacement:
                return Response({"detail": "No replacement questions available."}, status=404)
                
            original_order = question.order
            question.order = replacement.order
            replacement.order = original_order
            
            question.save(update_fields=['order'])
            replacement.save(update_fields=['order'])
            
            attempt.lifeline_switch_used = True
            attempt.save()
            
            choices = list(replacement.choices.all())
            return Response({
                "lifeline": "switch",
                "question": {
                    "id": replacement.id,
                    "text": replacement.text,
                    "category": replacement.category,
                    "order": original_order,
                    "choices": [{"id": c.id, "text": c.text} for c in choices]
                }
            })
            
        elif lifeline == 'expert':
            if attempt.lifeline_expert_used:
                return Response({"detail": "Ask the Expert lifeline already used."}, status=400)
                
            attempt.lifeline_expert_used = True
            attempt.lifeline_request_status = 'approved'
            attempt.pending_lifeline_type = 'expert'
            attempt.approved_lifeline_data = {
                "step": "select_expert"
            }
            attempt.save()
            
            return Response({
                "lifeline": "expert",
                "attempt": HotseatAttemptSerializer(attempt).data
            })


class HotseatWalkAwayView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        attempt.pending_lifeline_type = 'walkaway'
        attempt.lifeline_request_status = 'requested'
        attempt.approved_lifeline_data = {}
        attempt.save()
            
        return Response({
            "requested": True,
            "final_points": attempt.score
        })


class HotseatLifelineRequestView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if not attempt.options_visible:
            return Response({"detail": "Lifelines can only be requested after choices are shown by the host."}, status=400)
            
        if attempt.current_question_index >= 14:
            return Response({"detail": "Lifelines are no longer available on the 15th question."}, status=400)
            
        lifeline = request.data.get('lifeline') or request.data.get('lifeline_type')
        category = request.data.get('category', '')
        
        if not lifeline or lifeline not in ['5050', 'poll', 'switch', 'expert']:
            return Response({"detail": "Invalid lifeline provided."}, status=400)
            
        if lifeline == '5050' and attempt.lifeline_5050_used:
            return Response({"detail": "50:50 lifeline already used."}, status=400)
        elif lifeline == 'poll' and attempt.lifeline_poll_used:
            return Response({"detail": "Audience Poll lifeline already used."}, status=400)
        elif lifeline == 'switch' and attempt.lifeline_switch_used:
            return Response({"detail": "Switch Question lifeline already used."}, status=400)
        elif lifeline == 'expert' and attempt.lifeline_expert_used:
            return Response({"detail": "Ask the Expert lifeline already used."}, status=400)
            
        attempt.pending_lifeline_type = lifeline
        attempt.pending_lifeline_switch_category = category
        attempt.lifeline_request_status = 'requested'
        attempt.approved_lifeline_data = {}
        attempt.save()
        
        return Response({
            "detail": f"Request for {lifeline} lifeline submitted to host.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class HotseatLifelineAcknowledgeView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player and request.user.role != 'admin':
            return Response({"detail": "You are not the active hotseat contestant or admin."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        
        attempt.lifeline_request_status = 'none'
        attempt.pending_lifeline_type = ''
        attempt.pending_lifeline_switch_category = ''
        attempt.approved_lifeline_data = {}
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Lifeline status acknowledged.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminApproveLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat contestant."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'requested':
            return Response({"detail": "No pending lifeline request to approve."}, status=400)
            
        lifeline = attempt.pending_lifeline_type
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        approved_data = {}
        
        if lifeline == '5050':
            if attempt.lifeline_5050_used:
                return Response({"detail": "50:50 lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            incorrect_choices = [c for c in choices if not c.is_correct]
            
            import random
            random.shuffle(incorrect_choices)
            eliminated = incorrect_choices[:2]
            
            attempt.lifeline_5050_used = True
            approved_data = {
                "eliminated_choice_ids": [c.id for c in eliminated]
            }
            
        elif lifeline == 'poll':
            if attempt.lifeline_poll_used:
                return Response({"detail": "Audience Poll lifeline already used."}, status=400)
                
            from django.utils import timezone
            attempt.lifeline_poll_used = True
            approved_data = {
                "poll_start_time": timezone.now().isoformat(),
                "votes_closed": False,
                "votes": {}
            }
            
        elif lifeline == 'switch':
            if attempt.lifeline_switch_used:
                return Response({"detail": "Switch Question lifeline already used."}, status=400)
            approved_data = {}
            
        elif lifeline == 'expert':
            if attempt.lifeline_expert_used:
                return Response({"detail": "Ask the Expert lifeline already used."}, status=400)
            attempt.lifeline_expert_used = True
            approved_data = {
                "step": "select_expert"
            }
            
        elif lifeline == 'walkaway':
            attempt.status = HotseatAttempt.Status.WALKED_AWAY
            attempt.completed_at = timezone.now()
            approved_data = {}
            if batch_num == 1:
                quiz.hotseat_score_1 = attempt.score
                quiz.hotseat_status_1 = "walked_away"
                quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
            elif batch_num == 2:
                quiz.hotseat_score_2 = attempt.score
                quiz.hotseat_status_2 = "walked_away"
                quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
            elif batch_num == 3:
                quiz.hotseat_score_3 = attempt.score
                quiz.hotseat_status_3 = "walked_away"
                quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])
                
        attempt.lifeline_request_status = 'approved'
        attempt.approved_lifeline_data = approved_data
        attempt.save()
        
        return Response({
            "detail": f"{lifeline} lifeline approved successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminRejectLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        
        if attempt.lifeline_request_status != 'requested':
            return Response({"detail": "No pending lifeline request to reject."}, status=400)
            
        attempt.lifeline_request_status = 'rejected'
        attempt.save()
        
        return Response({
            "detail": "Lifeline request rejected successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminShowOptionsView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.options_visible = True
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Choices successfully revealed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminPauseTimerView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.timer_is_paused = True
        attempt.save()
        
        return Response({
            "detail": "Timer paused.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminResumeTimerView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Timer resumed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminNextQuestionReadyView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.showing_question = True
        attempt.options_visible = False
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Next question pushed to contestant.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminTriggerIntroView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.show_intro = True
        attempt.save()
        
        return Response({
            "detail": "KBC Intro playback triggered successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminCompleteIntroView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.show_intro = False
        attempt.intro_played = True
        attempt.save()
        
        return Response({
            "detail": "KBC Intro playback completed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class StudentSwitchCategoryListView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        categories = quiz.switch_categories.all().select_related('question')
        categories = [c for c in categories if c.question]
        
        data = []
        for c in categories:
            img_url = c.image.url if c.image else None
            if img_url and request:
                img_url = request.build_absolute_uri(img_url)
            data.append({
                "id": c.id,
                "name": c.name,
                "image": img_url
            })
        return Response(data)


class HotseatSelectSwitchCategoryView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'approved' or attempt.pending_lifeline_type != 'switch':
            return Response({"detail": "Switch Question lifeline has not been approved by the host."}, status=400)
            
        category_id = request.data.get('category_id')
        if not category_id:
            return Response({"detail": "Category ID is required."}, status=400)
            
        switch_cat = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        if not switch_cat.question:
            return Response({"detail": "No question configured for this category."}, status=400)
            
        with transaction.atomic():
            attempt.pending_lifeline_switch_category = f"{switch_cat.id}:{switch_cat.name}"
            attempt.save()
            
        return Response({
            "selected": True,
            "category_name": switch_cat.name,
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminConfirmSwitchLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat contestant."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'approved' or attempt.pending_lifeline_type != 'switch':
            return Response({"detail": "Switch Question lifeline is not active."}, status=400)
            
        selected_cat_str = attempt.pending_lifeline_switch_category
        if not selected_cat_str or ":" not in selected_cat_str:
            return Response({"detail": "Contestant has not selected a switch category yet."}, status=400)
            
        try:
            category_id = int(selected_cat_str.split(":", 1)[0])
        except Exception:
            return Response({"detail": "Invalid selected category format."}, status=400)
            
        switch_cat = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        replacement_question = switch_cat.question
        if not replacement_question:
            return Response({"detail": "No question configured for this category."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        original_question = questions[attempt.current_question_index]
        
        with transaction.atomic():
            original_order = original_question.order
            
            replacement_question.question_type = q_type
            replacement_question.order = original_order
            replacement_question.save()
            
            original_question.question_type = Question.QuestionType.SWITCH
            original_question.order = 999
            original_question.save()
            
            switch_cat.question = None
            switch_cat.save()
            
            attempt.lifeline_switch_used = True
            attempt.current_question_switched = True
            attempt.lifeline_request_status = 'none'
            attempt.pending_lifeline_type = ''
            attempt.pending_lifeline_switch_category = ''
            attempt.showing_question = True
            attempt.options_visible = False
            attempt.timer_is_paused = False
            attempt.save()
            
        return Response({
            "switched": True,
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class SystemPreferencesView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        prefs = SystemPreferences.get_solo()
        serializer = SystemPreferencesSerializer(prefs)
        return Response(serializer.data)

    def post(self, request):
        prefs = SystemPreferences.get_solo()
        serializer = SystemPreferencesSerializer(prefs, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SpectatorVoteView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        stage = quiz.current_stage
        active_hotseat_player = None
        active_batch = None
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            active_hotseat_player = quiz.hotseat_player_1
            active_batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            active_hotseat_player = quiz.hotseat_player_2
            active_batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            active_hotseat_player = quiz.hotseat_player_3
            active_batch = 3
            
        if not active_hotseat_player or not active_batch:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        attempt = HotseatAttempt.objects.filter(quiz=quiz, student=active_hotseat_player, batch_number=active_batch).first()
        if not attempt or attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "No active hotseat attempt running."}, status=400)
            
        if attempt.pending_lifeline_type != 'poll' or attempt.lifeline_request_status != 'approved':
            return Response({"detail": "Audience Poll lifeline is not active."}, status=400)
            
        from django.utils.dateparse import parse_datetime
        from django.utils import timezone
        
        approved_data = attempt.approved_lifeline_data or {}
        poll_start_time_str = approved_data.get('poll_start_time')
        votes_closed = approved_data.get('votes_closed', False)
        
        if not poll_start_time_str or votes_closed:
            return Response({"detail": "Audience Poll has closed."}, status=400)
            
        poll_start_time = parse_datetime(poll_start_time_str)
        if not poll_start_time:
            return Response({"detail": "Invalid poll start time."}, status=400)
            
        elapsed = (timezone.now() - poll_start_time).total_seconds()
        if elapsed >= 15.0:
            approved_data['votes_closed'] = True
            attempt.approved_lifeline_data = approved_data
            attempt.save(update_fields=['approved_lifeline_data'])
            return Response({"detail": "Audience Poll has closed."}, status=400)
            
        from quizzes.models import Question, Choice, SpectatorPollVote
        if active_batch == 1:
            q_type = Question.QuestionType.HOTSEAT_1
        elif active_batch == 2:
            q_type = Question.QuestionType.HOTSEAT_2
        else:
            q_type = Question.QuestionType.HOTSEAT_3
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "No active question found."}, status=400)
            
        question = questions[attempt.current_question_index]
        
        if SpectatorPollVote.objects.filter(attempt=attempt, student=request.user, question=question).exists():
            return Response({"detail": "You have already voted for this question."}, status=400)
            
        choice_id = request.data.get('choice_id')
        if not choice_id:
            return Response({"detail": "Choice ID is required."}, status=400)
            
        choice = Choice.objects.filter(question=question, id=choice_id).first()
        if not choice:
            return Response({"detail": "Invalid choice selected for this question."}, status=400)
            
        SpectatorPollVote.objects.create(
            attempt=attempt,
            student=request.user,
            question=question,
            choice=choice
        )
        
        return Response({"success": True, "detail": "Vote submitted successfully!"})


class QuizDetailedReportView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        user = request.user
        quiz = get_object_or_404(Quiz, pk=pk)
        if not getattr(user, 'is_super_admin', False) and quiz.created_by != user and quiz.host != user:
            return Response({"detail": "You do not have access to this quiz report."}, status=403)
        
        from quizzes.reports import generate_quiz_pdf_report
        try:
            pdf_data = generate_quiz_pdf_report(quiz.id)
        except Exception as e:
            return Response({"detail": f"Error generating report: {str(e)}"}, status=500)
            
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="kbc_report_{quiz.id}.pdf"'
        return response


class BuzzerInitView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state, created = BuzzerState.objects.get_or_create(quiz=quiz)
        if not state.buzzer_mappings:
            state.buzzer_mappings = {str(i): {"name": f"Podium {i}", "score": 0} for i in range(1, 16)}
            state.save(update_fields=['buzzer_mappings'])
        if not state.current_question:
            q = Question.objects.filter(quiz=quiz, question_type=Question.QuestionType.BUZZER).order_by('order', 'id').first()
            if q:
                state.current_question = q
                state.save(update_fields=['current_question'])
        return Response(BuzzerStateSerializer(state).data)


class BuzzerNextQuestionView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        questions = list(Question.objects.filter(quiz=quiz, question_type=Question.QuestionType.BUZZER).order_by('order', 'id'))
        if not questions:
            return Response({"detail": "No Buzzer Round questions found."}, status=400)
        
        curr_q = state.current_question
        next_q = None
        if not curr_q:
            next_q = questions[0]
        else:
            try:
                curr_idx = questions.index(curr_q)
                if curr_idx + 1 < len(questions):
                    next_q = questions[curr_idx + 1]
                else:
                    return Response({"detail": "You are already at the last Buzzer question."}, status=400)
            except ValueError:
                next_q = questions[0]
        
        state.current_question = next_q
        state.buzzers_locked = True
        state.is_timer_running = False
        state.timer_started_at = None
        state.timer_paused_at = None
        state.incorrect_buzzers = []
        state.active_buzzer_id = None
        state.options_visible = False
        state.answer_visible = False
        state.save()
        
        BuzzerPress.objects.filter(quiz=quiz, question=next_q).delete()
        return Response(BuzzerStateSerializer(state).data)


class BuzzerReleaseView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        if not state.current_question:
            return Response({"detail": "No active question selected."}, status=400)
        
        state.buzzers_locked = False
        state.active_buzzer_id = None
        state.is_timer_running = False
        state.timer_started_at = None
        state.timer_paused_at = None
        state.save()
        
        BuzzerPress.objects.filter(quiz=quiz, question=state.current_question).delete()
        return Response(BuzzerStateSerializer(state).data)


class BuzzerAnswerCorrectView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        active_buzzer = state.active_buzzer_id
        if not active_buzzer:
            return Response({"detail": "No active buzzer is currently selected to answer."}, status=400)
            
        mappings = state.buzzer_mappings or {}
        if active_buzzer in mappings:
            points = state.current_question.marks if state.current_question else 1
            mappings[active_buzzer]["score"] = mappings[active_buzzer].get("score", 0) + points
            state.buzzer_mappings = mappings
            
        state.buzzers_locked = True
        state.is_timer_running = False
        state.options_visible = True
        state.answer_visible = True
        state.active_buzzer_id = None  # Clear so the answering panel hides
        state.save()
        
        return Response(BuzzerStateSerializer(state).data)


class BuzzerAnswerIncorrectView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        active_buzzer = state.active_buzzer_id
        if not active_buzzer:
            return Response({"detail": "No active buzzer is currently selected."}, status=400)
            
        incorrect = state.incorrect_buzzers or []
        if active_buzzer not in incorrect:
            incorrect.append(active_buzzer)
            state.incorrect_buzzers = incorrect
            
        state.active_buzzer_id = None
        state.is_timer_running = False
        state.timer_started_at = None
        state.timer_paused_at = None
        
        # Check if all active teams have answered incorrectly
        if len(state.incorrect_buzzers) >= (state.buzzer_count or 15):
            state.options_visible = True
            state.answer_visible = True
            state.buzzers_locked = True
        else:
            state.buzzers_locked = False
            
        state.save()
        return Response(BuzzerStateSerializer(state).data)


class BuzzerResetView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        state.buzzers_locked = True
        state.is_timer_running = False
        state.timer_started_at = None
        state.timer_paused_at = None
        state.incorrect_buzzers = []
        state.active_buzzer_id = None
        state.options_visible = False
        state.answer_visible = False
        state.save()
        
        BuzzerPress.objects.filter(quiz=quiz, question=state.current_question).delete()
        return Response(BuzzerStateSerializer(state).data)


class BuzzerUpdateMappingsView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        mappings = request.data.get('mappings')
        timer_limit = request.data.get('timer_limit')
        buzzer_count = request.data.get('buzzer_count')
        
        if mappings is not None:
            state.buzzer_mappings = mappings
        if timer_limit is not None:
            state.answer_timer_limit = int(timer_limit)
        if buzzer_count is not None:
            state.buzzer_count = max(1, min(30, int(buzzer_count)))
            
        state.save()
        return Response(BuzzerStateSerializer(state).data)


class BuzzerRevealOptionsView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        state.options_visible = not state.options_visible
        state.save(update_fields=['options_visible'])
        return Response(BuzzerStateSerializer(state).data)


class BuzzerRevealAnswerView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        state = get_object_or_404(BuzzerState, quiz=quiz)
        state.answer_visible = not state.answer_visible
        state.save(update_fields=['answer_visible'])
        return Response(BuzzerStateSerializer(state).data)


class PressBuzzerView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if quiz.current_stage != Quiz.Stage.BUZZER_ROUND:
            return Response({"detail": "Quiz is not in Buzzer Round stage."}, status=400)
            
        state, _ = BuzzerState.objects.get_or_create(quiz=quiz)
            
        if not state.current_question:
            return Response({"detail": "No active question is selected."}, status=400)
            
        buzzer_id = str(request.data.get('buzzer_id', '')).strip()
        if not buzzer_id:
            return Response({"detail": "Buzzer ID is required."}, status=400)
            
        if buzzer_id in (state.incorrect_buzzers or []):
            return Response({"detail": "This buzzer has already answered incorrectly and is blocked."}, status=400)
            
        if BuzzerPress.objects.filter(quiz=quiz, question=state.current_question, buzzer_id=buzzer_id).exists():
            return Response({"detail": "Buzzer already registered."}, status=400)
            
        BuzzerPress.objects.create(
            quiz=quiz,
            question=state.current_question,
            buzzer_id=buzzer_id,
            time_taken_seconds=0.0
        )
        
        if not state.active_buzzer_id:
            state.active_buzzer_id = buzzer_id
            state.buzzers_locked = True
            state.is_timer_running = True
            state.timer_started_at = timezone.now()
            state.timer_paused_at = None
            state.save()
            
        return Response({
            "success": True,
            "detail": "Buzzer press registered successfully!",
            "active_buzzer_id": state.active_buzzer_id
        })


class StudentExpertListView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        experts = quiz.experts.all()
        data = []
        for e in experts:
            photo_url = e.photo.url if e.photo else None
            if photo_url and request:
                photo_url = request.build_absolute_uri(photo_url)
            data.append({
                "id": e.id,
                "name": e.name,
                "designation": e.designation,
                "photo": photo_url
            })
        return Response(data)


class HotseatSelectExpertView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'approved' or attempt.pending_lifeline_type != 'expert':
            return Response({"detail": "Ask the Expert lifeline is not approved."}, status=400)
            
        expert_id = request.data.get('expert_id')
        if not expert_id:
            return Response({"detail": "expert_id is required."}, status=400)
            
        expert = get_object_or_404(Expert, quiz=quiz, id=expert_id)
        
        photo_url = expert.photo.url if expert.photo else None
        if photo_url and request:
            photo_url = request.build_absolute_uri(photo_url)
            
        prefs = SystemPreferences.get_solo()
        attempt.approved_lifeline_data = {
            "step": "timer",
            "selected_expert": {
                "id": expert.id,
                "name": expert.name,
                "designation": expert.designation,
                "photo": photo_url
            },
            "timer_started_at": timezone.now().isoformat(),
            "timer_duration": getattr(prefs, 'expert_timer_limit', 30)
        }
        # Pause the main countdown timer
        attempt.timer_is_paused = True
        attempt.save()
        
        return Response({
            "detail": "Expert selected and timer started.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })



class AdminTeamManagementView(APIView):
    """
    Admin-only endpoint to list, create, update, and delete teams for a quiz.
    Allows creating teams from 1–4 registered students (no leader restriction).
    """
    permission_classes = [IsAdminUser]

    def _get_quiz(self, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(self.request.user, quiz):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You do not have permission to manage this quiz.")
        return quiz

    def _resolve_member(self, identifier, quiz, idx, current_team_id=None):
        """Resolve a user by email or id, validate quiz registration."""
        User = get_user_model()
        try:
            if isinstance(identifier, int) or str(identifier).isdigit():
                user = User.objects.get(pk=int(identifier))
            else:
                user = User.objects.get(email__iexact=str(identifier).strip())
        except User.DoesNotExist:
            raise ValidationError({"detail": f"Member {idx} not found in the system."})

        if not QuizRegistration.objects.filter(quiz=quiz, student=user).exists():
            raise ValidationError({"detail": f"Member {idx} ({user.full_name}) is not registered for this quiz."})

        # Check if already in another team
        from quizzes.models import Team as TeamModel
        other_teams = TeamModel.objects.filter(quiz=quiz).filter(Q(leader=user) | Q(members=user))
        if current_team_id:
            other_teams = other_teams.exclude(pk=current_team_id)
        if other_teams.exists():
            raise ValidationError({"detail": f"Member {idx} ({user.full_name}) is already in another team for this quiz."})

        return user

    def get(self, request, pk):
        quiz = self._get_quiz(pk)
        from quizzes.models import Team as TeamModel
        from quizzes.serializers import TeamSerializer
        teams = TeamModel.objects.filter(quiz=quiz).prefetch_related('members').select_related('leader')

        # Attach which podium each team is assigned to
        buzzer_state = getattr(quiz, 'buzzer_state', None)
        mappings = buzzer_state.buzzer_mappings if buzzer_state else {}
        team_to_podium = {}
        for podium_id, mapping_data in mappings.items():
            tid = str(mapping_data.get('team_id', ''))
            if tid:
                team_to_podium[tid] = int(podium_id)

        data = []
        for t in teams:
            serialized = TeamSerializer(t).data
            serialized['podium'] = team_to_podium.get(str(t.id))
            data.append(serialized)

        return Response(data)

    def post(self, request, pk):
        quiz = self._get_quiz(pk)
        from quizzes.models import Team as TeamModel

        name = request.data.get('name', '').strip()
        if not name:
            return Response({"detail": "Team name is required."}, status=400)

        if TeamModel.objects.filter(quiz=quiz, name__iexact=name).exists():
            return Response({"detail": f"A team named '{name}' already exists in this quiz."}, status=400)

        # Collect member identifiers (email or id). At least 1 required, max 4.
        member_fields = ['member1', 'member2', 'member3', 'member4']
        raw_members = [request.data.get(f) for f in member_fields]
        raw_members = [m for m in raw_members if m]  # drop blanks

        if not raw_members:
            return Response({"detail": "At least 1 member is required."}, status=400)
        if len(raw_members) > 4:
            return Response({"detail": "A team can have at most 4 members."}, status=400)

        # Check uniqueness across members
        seen = set()
        resolved_users = []
        for idx, identifier in enumerate(raw_members, 1):
            user = self._resolve_member(identifier, quiz, idx)
            uid = user.pk
            if uid in seen:
                return Response({"detail": f"Duplicate member detected: {user.full_name}."}, status=400)
            seen.add(uid)
            resolved_users.append(user)

        # Create team — leader is the first member
        leader = resolved_users[0]
        name_fields = {}
        email_fields = {}
        for i, u in enumerate(resolved_users, 1):
            name_fields[f'member{i}_name'] = u.full_name
            email_fields[f'member{i}_email'] = u.email

        with transaction.atomic():
            from quizzes.models import Team as TeamModel
            team = TeamModel.objects.create(
                quiz=quiz,
                name=name,
                leader=leader,
                **name_fields,
                **email_fields,
            )
            # Members M2M excludes leader
            non_leaders = resolved_users[1:]
            team.members.set(non_leaders)

        from quizzes.serializers import TeamSerializer
        return Response(TeamSerializer(team).data, status=201)

    def patch(self, request, pk, team_id):
        quiz = self._get_quiz(pk)
        from quizzes.models import Team as TeamModel
        team = get_object_or_404(TeamModel, pk=team_id, quiz=quiz)

        name = request.data.get('name', '').strip() or team.name
        if name != team.name and TeamModel.objects.filter(quiz=quiz, name__iexact=name).exclude(pk=team.pk).exists():
            return Response({"detail": f"A team named '{name}' already exists in this quiz."}, status=400)

        member_fields = ['member1', 'member2', 'member3', 'member4']
        raw_members = [request.data.get(f) for f in member_fields]
        raw_members = [m for m in raw_members if m]

        if not raw_members:
            # No members provided — just update name
            team.name = name
            team.save(update_fields=['name'])
            from quizzes.serializers import TeamSerializer
            return Response(TeamSerializer(team).data)

        if len(raw_members) > 4:
            return Response({"detail": "A team can have at most 4 members."}, status=400)

        seen = set()
        resolved_users = []
        for idx, identifier in enumerate(raw_members, 1):
            user = self._resolve_member(identifier, quiz, idx, current_team_id=team.pk)
            uid = user.pk
            if uid in seen:
                return Response({"detail": f"Duplicate member detected: {user.full_name}."}, status=400)
            seen.add(uid)
            resolved_users.append(user)

        leader = resolved_users[0]
        name_fields = {f'member{i}_name': '' for i in range(1, 5)}
        email_fields = {f'member{i}_email': '' for i in range(1, 5)}
        for i, u in enumerate(resolved_users, 1):
            name_fields[f'member{i}_name'] = u.full_name
            email_fields[f'member{i}_email'] = u.email

        with transaction.atomic():
            team.name = name
            team.leader = leader
            for field, val in {**name_fields, **email_fields}.items():
                setattr(team, field, val)
            team.save()
            team.members.set(resolved_users[1:])

        from quizzes.serializers import TeamSerializer
        return Response(TeamSerializer(team).data)

    def delete(self, request, pk, team_id):
        quiz = self._get_quiz(pk)
        from quizzes.models import Team as TeamModel
        team = get_object_or_404(TeamModel, pk=team_id, quiz=quiz)
        team.delete()
        return Response({"detail": "Team deleted successfully."})


class UnteamedPlayersView(APIView):
    """
    Admin-only endpoint. Returns all students registered for the quiz
    who are NOT yet part of any team for that quiz.
    """
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        if not check_admin_quiz_write_access(request.user, quiz):
            return Response({"detail": "You do not have permission to manage this quiz."}, status=status.HTTP_403_FORBIDDEN)
        from quizzes.models import Team as TeamModel

        # All users in any team for this quiz
        teamed_ids = set()
        for team in TeamModel.objects.filter(quiz=quiz).prefetch_related('members').select_related('leader'):
            teamed_ids.add(team.leader_id)
            for m in team.members.all():
                teamed_ids.add(m.pk)

        registrations = QuizRegistration.objects.filter(quiz=quiz).select_related('student').exclude(
            student_id__in=teamed_ids
        )

        data = []
        for r in registrations:
            s = r.student
            data.append({
                "id": s.pk,
                "full_name": s.full_name,
                "email": s.email,
            })

        return Response(data)



