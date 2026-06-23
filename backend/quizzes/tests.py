from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from quizzes.models import Quiz, QuizRegistration, Question, Choice
from users.models import School, Program, Branch, StudentProfile
import io
import openpyxl

User = get_user_model()

class QuizEnrollmentTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        
        # Create school program branch
        self.school = School.objects.create(school_name="Engineering School", school_code="SOET")
        self.program = Program.objects.create(school=self.school, program_name="Computer Science", program_code="CSE_PROG")
        self.branch = Branch.objects.create(program=self.program, branch_name="CSE Branch", branch_code="CSE")
        
        # Create admin user
        self.admin_user = User.objects.create_superuser(
            email="admin@quizverse.edu",
            password="adminpassword123",
            full_name="System Admin",
            college_id="ADMIN001"
        )
        
        # Authenticate
        self.client.force_authenticate(user=self.admin_user)
        
        # Create active quiz
        self.quiz = Quiz.objects.create(
            title="KBC Arena Live Testing Quiz",
            description="Testing live event",
            status=Quiz.Status.REGISTRATION_OPEN,
            event_password="KBC123",
            created_by=self.admin_user
        )

    def test_manual_enroll_student(self):
        url = f"/api/quizzes/admin/{self.quiz.id}/enroll_student_manual/"
        payload = {
            "email": "contestant@quizverse.edu",
            "full_name": "Contestant One",
            "college_id": "ST001",
            "payment_status": "paid"
        }
        
        response = self.client.post(url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("Successfully enrolled", response.data["detail"])
        
        # Verify database creation
        self.assertTrue(User.objects.filter(email="contestant@quizverse.edu").exists())
        student = User.objects.get(email="contestant@quizverse.edu")
        self.assertEqual(student.full_name, "Contestant One")
        self.assertEqual(student.college_id, "ST001")
        
        # Verify student profile created automatically
        self.assertTrue(hasattr(student, "student_profile"))
        
        # Verify registration created
        self.assertTrue(QuizRegistration.objects.filter(student=student, quiz=self.quiz).exists())
        reg = QuizRegistration.objects.get(student=student, quiz=self.quiz)
        self.assertEqual(reg.payment_status, "paid")
        self.assertEqual(reg.player_id, "PLAYER 001")

    def test_manual_enroll_duplicate_returns_ok_with_updated_status(self):
        # First enroll
        url = f"/api/quizzes/admin/{self.quiz.id}/enroll_student_manual/"
        payload = {
            "email": "contestant@quizverse.edu",
            "full_name": "Contestant One",
            "college_id": "ST001",
            "payment_status": "pending"
        }
        self.client.post(url, payload, format="json")
        
        # Re-enroll with status paid
        payload["payment_status"] = "paid"
        response = self.client.post(url, payload, format="json")
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("already registered", response.data["detail"])
        
        # Check DB payment_status updated
        reg = QuizRegistration.objects.get(student__email="contestant@quizverse.edu", quiz=self.quiz)
        self.assertEqual(reg.payment_status, "paid")

    def test_download_enrollment_template(self):
        url = "/api/quizzes/admin/download_enrollment_template/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Verify xlsx contents
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Student Enrollment Template")
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ['Full Name', 'Email', 'Roll Number', 'Payment Status (paid/pending)'])

    def test_download_buzzer_template(self):
        url = "/api/quizzes/admin/download_template/?type=buzzer"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Verify xlsx contents
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Buzzer Round Template")
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, [
            'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
            'Correct Option (A/B/C/D)', 'Marks', 
            'Question Type (buzzer)', 'Category', 'Trivia'
        ])

    def test_bulk_enroll_students_csv(self):
        url = f"/api/quizzes/admin/{self.quiz.id}/bulk_enroll_students/"
        csv_content = (
            "Full Name,Email,Roll Number,Payment Status (paid/pending)\n"
            "Bulk Student 1,bulk1@quizverse.edu,ST_BULK1,paid\n"
            "Bulk Student 2,bulk2@quizverse.edu,ST_BULK2,pending\n"
        )
        csv_file = io.BytesIO(csv_content.encode("utf-8"))
        csv_file.name = "students.csv"
        
        response = self.client.post(url, {"file": csv_file}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("Successfully enrolled 2 students", response.data["detail"])
        
        # Verify databases
        self.assertEqual(User.objects.filter(email__contains="bulk").count(), 2)
        u1 = User.objects.get(email="bulk1@quizverse.edu")
        u2 = User.objects.get(email="bulk2@quizverse.edu")
        self.assertEqual(u1.college_id, "ST_BULK1")
        self.assertEqual(u2.college_id, "ST_BULK2")
        self.assertEqual(u1.roll_number, "ST_BULK1")
        self.assertEqual(u2.roll_number, "ST_BULK2")
        
        reg1 = QuizRegistration.objects.get(student=u1, quiz=self.quiz)
        reg2 = QuizRegistration.objects.get(student=u2, quiz=self.quiz)
        self.assertEqual(reg1.payment_status, "paid")
        self.assertEqual(reg2.payment_status, "pending")


class QuizBuzzerTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        
        # Create school program branch
        self.school = School.objects.create(school_name="Engineering School", school_code="SOET")
        self.program = Program.objects.create(school=self.school, program_name="Computer Science", program_code="CSE_PROG")
        self.branch = Branch.objects.create(program=self.program, branch_name="CSE Branch", branch_code="CSE")
        
        # Create admin user
        self.admin_user = User.objects.create_superuser(
            email="admin2@quizverse.edu",
            password="adminpassword123",
            full_name="System Admin",
            college_id="ADMIN002"
        )
        self.admin_user.role = "admin"
        self.admin_user.save()
        
        # Create student users
        self.students = []
        for i in range(5):
            student = User.objects.create_user(
                email=f"student{i}@quizverse.edu",
                password="studentpassword123",
                full_name=f"Student {i}",
                college_id=f"ST{100 + i}",
                role="student"
            )
            StudentProfile.objects.create(
                user=student,
                school=self.school,
                program=self.program,
                branch=self.branch,
                year=StudentProfile.Year.FIRST
            )
            self.students.append(student)

        # Create active quiz in Buzzer Round stage
        self.quiz = Quiz.objects.create(
            title="Buzzer Quiz",
            description="Buzzer round testing",
            status=Quiz.Status.REGISTRATION_OPEN,
            event_password="BUZZ123",
            created_by=self.admin_user,
            current_stage=Quiz.Stage.BUZZER_ROUND
        )

        # Register students
        for student in self.students:
            QuizRegistration.objects.create(
                student=student,
                quiz=self.quiz,
                payment_status="paid"
            )

        # Create a buzzer round question
        self.question = Question.objects.create(
            quiz=self.quiz,
            text="Sample Buzzer Question",
            question_type=Question.QuestionType.BUZZER,
            marks=10,
            order=1
        )
        
        Choice.objects.create(question=self.question, text="Choice A", is_correct=True)
        Choice.objects.create(question=self.question, text="Choice B", is_correct=False)
        Choice.objects.create(question=self.question, text="Choice C", is_correct=False)
        Choice.objects.create(question=self.question, text="Choice D", is_correct=False)

    def test_team_size_limit(self):
        # Authenticate first student
        self.client.force_authenticate(user=self.students[0])
        
        # Create team with exactly 4 members
        create_url = "/api/quizzes/teams/"
        payload = {
            "quiz": self.quiz.id,
            "name": "Buzzer Team",
            "member1_email": self.students[0].email,
            "member2_email": self.students[1].email,
            "member3_email": self.students[2].email,
            "member4_email": self.students[3].email,
        }
        response = self.client.post(create_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["member1_name"], self.students[0].full_name)
        self.assertEqual(response.data["member2_name"], self.students[1].full_name)
        self.assertEqual(response.data["member3_name"], self.students[2].full_name)
        self.assertEqual(response.data["member4_name"], self.students[3].full_name)
        team_id = response.data["id"]
        
        # Try to join 5th student - should fail because team is already size 4
        self.client.force_authenticate(user=self.students[4])
        join_url = f"/api/quizzes/teams/{team_id}/join/"
        response = self.client.post(join_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("maximum size of 4 members", response.data["detail"])

    def test_team_creation_missing_emails(self):
        self.client.force_authenticate(user=self.students[0])
        create_url = "/api/quizzes/teams/"
        payload = {
            "quiz": self.quiz.id,
            "name": "Invalid Team",
            "member1_email": self.students[0].email,
            "member2_email": self.students[1].email,
        }
        response = self.client.post(create_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("email is required", response.data["detail"])

    def test_team_creation_unregistered_member(self):
        unregistered_student = User.objects.create_user(
            email="unregistered@quizverse.edu",
            password="studentpassword123",
            full_name="Unregistered Student",
            college_id="ST999",
            role="student"
        )
        StudentProfile.objects.create(
            user=unregistered_student,
            school=self.school,
            program=self.program,
            branch=self.branch,
            year=StudentProfile.Year.FIRST
        )
        
        self.client.force_authenticate(user=self.students[0])
        create_url = "/api/quizzes/teams/"
        payload = {
            "quiz": self.quiz.id,
            "name": "Invalid Team",
            "member1_email": self.students[0].email,
            "member2_email": self.students[1].email,
            "member3_email": self.students[2].email,
            "member4_email": unregistered_student.email,
        }
        response = self.client.post(create_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("is not registered for this quiz", response.data["detail"])

    def test_team_creation_duplicate_members(self):
        self.client.force_authenticate(user=self.students[0])
        create_url = "/api/quizzes/teams/"
        payload = {
            "quiz": self.quiz.id,
            "name": "Invalid Team",
            "member1_email": self.students[0].email,
            "member2_email": self.students[1].email,
            "member3_email": self.students[2].email,
            "member4_email": self.students[2].email,
        }
        response = self.client.post(create_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("unique emails", response.data["detail"])

    def test_buzzer_round_mechanics(self):
        # 1. Initialize buzzer round
        self.client.force_authenticate(user=self.admin_user)
        init_url = f"/api/quizzes/admin/{self.quiz.id}/buzzer_init/"
        response = self.client.post(init_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify initial state
        state_data = response.data
        self.assertEqual(state_data["current_question"]["id"], self.question.id)
        self.assertTrue(state_data["buzzers_locked"])
        self.assertIsNone(state_data["active_buzzer_id"])
        
        # 2. Press buzzer (succeeds even though buzzers are technically in initial locked state)
        self.client.force_authenticate(user=self.students[0])
        press_url = f"/api/quizzes/{self.quiz.id}/press-buzzer/"
        response = self.client.post(press_url, {"buzzer_id": "1"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["success"])
        self.assertEqual(response.data["active_buzzer_id"], "1")
        
        # 3. Release buzzers (unlock/reset active buzzer)
        self.client.force_authenticate(user=self.admin_user)
        release_url = f"/api/quizzes/admin/{self.quiz.id}/buzzer_release/"
        response = self.client.post(release_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["buzzers_locked"])
        
        # 4. Check live state - timer is not running, active_buzzer_id is None
        live_state_url = f"/api/quizzes/{self.quiz.id}/live-state/"
        self.client.force_authenticate(user=self.students[0])
        response = self.client.get(live_state_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        buzzer_state = response.data["buzzer_state"]
        self.assertIsNone(buzzer_state["active_buzzer_id"])
        self.assertFalse(buzzer_state["is_timer_running"])
        
        # 5. Press buzzer 1 again (succeeds because release deleted the previous press)
        self.client.force_authenticate(user=self.students[0])
        response = self.client.post(press_url, {"buzzer_id": "1"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["active_buzzer_id"], "1")

        # 6. Try to press buzzer 1 yet again (fails because it's already registered now)
        response = self.client.post(press_url, {"buzzer_id": "1"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("already registered", response.data["detail"])
        
        # 7. Press another buzzer 2 (succeeds immediately because locks are bypassed)
        self.client.force_authenticate(user=self.students[1])
        response = self.client.post(press_url, {"buzzer_id": "2"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["active_buzzer_id"], "1")
        
        # 8. Admin marks buzzer 1 answer as incorrect (adds to incorrect list, clears active)
        self.client.force_authenticate(user=self.admin_user)
        incorrect_url = f"/api/quizzes/admin/{self.quiz.id}/buzzer_answer_incorrect/"
        response = self.client.post(incorrect_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNone(response.data["active_buzzer_id"])
        self.assertFalse(response.data["is_timer_running"])
        self.assertIn("1", response.data["incorrect_buzzers"])
        
        # 9. Try to press buzzer 1 again (should be blocked since incorrect)
        self.client.force_authenticate(user=self.students[0])
        response = self.client.post(press_url, {"buzzer_id": "1"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("blocked", response.data["detail"])
        
        # 10. Press buzzer 3 (valid new buzzer)
        self.client.force_authenticate(user=self.students[2])
        response = self.client.post(press_url, {"buzzer_id": "3"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["active_buzzer_id"], "3")
        
        # 11. Admin marks buzzer 3 answer as correct (awards points)
        self.client.force_authenticate(user=self.admin_user)
        correct_url = f"/api/quizzes/admin/{self.quiz.id}/buzzer_answer_correct/"
        response = self.client.post(correct_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["answer_visible"])
        self.assertEqual(response.data["buzzer_mappings"]["3"]["score"], 10)
        
        # 12. Update mappings and timer limit
        update_url = f"/api/quizzes/admin/{self.quiz.id}/buzzer_update_mappings/"
        new_mappings = response.data["buzzer_mappings"]
        new_mappings["3"]["name"] = "Team Super"
        response = self.client.post(update_url, {"mappings": new_mappings, "timer_limit": 25}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["buzzer_mappings"]["3"]["name"], "Team Super")
        self.assertEqual(response.data["answer_timer_limit"], 25)

    def test_registered_players_endpoint(self):
        # Student 1 is registered, so they should be able to view registered players
        self.client.force_authenticate(user=self.students[0])
        url = f"/api/quizzes/{self.quiz.id}/registered-players/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Should return all 5 registered students in self.students
        self.assertEqual(len(response.data), 5)
        emails = [item["email"] for item in response.data]
        self.assertIn(self.students[0].email, emails)
        self.assertIn(self.students[1].email, emails)

        # A non-registered user should be forbidden
        unregistered = User.objects.create_user(
            email="nonreg@quizverse.edu",
            password="password",
            full_name="Non Registered",
            college_id="NONREG1",
            role="student"
        )
        self.client.force_authenticate(user=unregistered)
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_fetching_teams(self):
        # Create a team first
        self.client.force_authenticate(user=self.students[0])
        create_url = "/api/quizzes/teams/"
        payload = {
            "quiz": self.quiz.id,
            "name": "Admin Test Team",
            "member1_email": self.students[0].email,
            "member2_email": self.students[1].email,
            "member3_email": self.students[2].email,
            "member4_email": self.students[3].email,
        }
        response = self.client.post(create_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        # Now admin tries to list teams
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get("/api/quizzes/teams/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["name"], "Admin Test Team")


class QuizExpertLifelineTests(TestCase):
    def setUp(self):
        from quizzes.models import Quiz, QuizRegistration, Question, Choice, Expert, HotseatAttempt, SystemPreferences
        from users.models import School, Program, Branch, StudentProfile
        self.client = APIClient()
        
        # Create school program branch
        self.school = School.objects.create(school_name="Engineering School", school_code="SOET")
        self.program = Program.objects.create(school=self.school, program_name="Computer Science", program_code="CSE_PROG")
        self.branch = Branch.objects.create(program=self.program, branch_name="CSE Branch", branch_code="CSE")
        
        # Create admin user
        self.admin_user = User.objects.create_superuser(
            email="admin@quizverse.edu",
            password="adminpassword123",
            full_name="System Admin",
            college_id="ADMIN001"
        )
        
        # Create student user
        self.student = User.objects.create_user(
            email="student@quizverse.edu",
            password="studentpassword123",
            full_name="Student One",
            college_id="ST001",
            role="student"
        )
        StudentProfile.objects.create(
            user=self.student,
            school=self.school,
            program=self.program,
            branch=self.branch,
            year=StudentProfile.Year.FIRST
        )
        
        # Create quiz
        self.quiz = Quiz.objects.create(
            title="KBC Expert Arena",
            description="Testing expert lifeline",
            status=Quiz.Status.REGISTRATION_OPEN,
            event_password="KBC123",
            created_by=self.admin_user,
            current_stage=Quiz.Stage.HOTSEAT_BATCH_1,
            hotseat_player_1=self.student
        )
        
        # Enroll student in quiz
        QuizRegistration.objects.create(
            student=self.student,
            quiz=self.quiz,
            payment_status="paid"
        )
        
        # Create a question for Hotseat Batch 1
        self.question = Question.objects.create(
            quiz=self.quiz,
            text="Which language is this?",
            question_type=Question.QuestionType.HOTSEAT_1,
            order=1,
            marks=100
        )
        self.choice1 = Choice.objects.create(question=self.question, text="Python", is_correct=True)
        self.choice2 = Choice.objects.create(question=self.question, text="C++", is_correct=False)

        # Create hotseat attempt
        self.attempt = HotseatAttempt.objects.create(
            quiz=self.quiz,
            student=self.student,
            batch_number=1,
            status=HotseatAttempt.Status.PLAYING,
            current_question_index=0,
            options_visible=True
        )

    def test_admin_expert_crud(self):
        # Authenticate admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Save expert (Create)
        url = f"/api/quizzes/admin/{self.quiz.id}/save_expert/"
        payload = {
            "name": "Dr. Vani Agarwal",
            "designation": "Assistant Professor"
        }
        response = self.client.post(url, payload, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["expert"]["name"], "Dr. Vani Agarwal")
        self.assertEqual(response.data["expert"]["designation"], "Assistant Professor")
        expert_id = response.data["expert"]["id"]
        
        # Get experts
        url_get = f"/api/quizzes/admin/{self.quiz.id}/get_experts/"
        response = self.client.get(url_get)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["name"], "Dr. Vani Agarwal")
        
        # Edit expert
        payload_edit = {
            "expert_id": expert_id,
            "name": "Dr. Vani Agarwal Edited",
            "designation": "Associate Professor"
        }
        response = self.client.post(url, payload_edit, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["expert"]["name"], "Dr. Vani Agarwal Edited")
        self.assertEqual(response.data["expert"]["designation"], "Associate Professor")
        
        # Delete expert
        url_delete = f"/api/quizzes/admin/{self.quiz.id}/delete_expert/"
        response = self.client.post(url_delete, {"expert_id": expert_id}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify deleted
        response = self.client.get(url_get)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 0)

    def test_expert_limit(self):
        from quizzes.models import Expert
        # Authenticate admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Create 5 experts
        for i in range(5):
            Expert.objects.create(
                quiz=self.quiz,
                name=f"Expert {i}",
                designation="Professor"
            )
            
        # Try to save a 6th expert - should fail
        url = f"/api/quizzes/admin/{self.quiz.id}/save_expert/"
        payload = {
            "name": "Expert 6",
            "designation": "Professor"
        }
        response = self.client.post(url, payload, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Maximum of 5 experts", response.data["detail"])

    def test_student_list_experts(self):
        from quizzes.models import Expert
        # Create an expert
        Expert.objects.create(
            quiz=self.quiz,
            name="Dr. Vani Agarwal",
            designation="Assistant Professor"
        )
        
        # Authenticate student
        self.client.force_authenticate(user=self.student)
        url = f"/api/quizzes/{self.quiz.id}/experts/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["name"], "Dr. Vani Agarwal")
        self.assertEqual(response.data[0]["designation"], "Assistant Professor")

    def test_lifeline_expert_flow(self):
        from quizzes.models import Expert, HotseatAttempt
        expert = Expert.objects.create(
            quiz=self.quiz,
            name="Dr. Vani Agarwal",
            designation="Assistant Professor"
        )
        
        # Step 1: Student requests expert lifeline
        self.client.force_authenticate(user=self.student)
        req_url = f"/api/quizzes/{self.quiz.id}/hotseat-lifeline-request/"
        response = self.client.post(req_url, {"lifeline": "expert"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify attempt state updated
        attempt = HotseatAttempt.objects.get(id=self.attempt.id)
        self.assertEqual(attempt.lifeline_request_status, 'requested')
        self.assertEqual(attempt.pending_lifeline_type, 'expert')
        
        # Step 2: Host approves the lifeline
        self.client.force_authenticate(user=self.admin_user)
        approve_url = f"/api/quizzes/admin/{self.quiz.id}/approve_lifeline/"
        response = self.client.post(approve_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify approved state
        attempt = HotseatAttempt.objects.get(id=self.attempt.id)
        self.assertEqual(attempt.lifeline_request_status, 'approved')
        self.assertEqual(attempt.approved_lifeline_data["step"], 'select_expert')
        
        # Step 3: Student selects the expert
        self.client.force_authenticate(user=self.student)
        select_url = f"/api/quizzes/{self.quiz.id}/hotseat-select-expert/"
        response = self.client.post(select_url, {"expert_id": expert.id}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify countdown timer starts and quiz timer is paused
        attempt = HotseatAttempt.objects.get(id=self.attempt.id)
        self.assertTrue(attempt.timer_is_paused)
        self.assertEqual(attempt.approved_lifeline_data["step"], 'timer')
        self.assertEqual(attempt.approved_lifeline_data["selected_expert"]["name"], "Dr. Vani Agarwal")
        self.assertEqual(attempt.approved_lifeline_data["timer_duration"], 30) # Default
        
        # Step 4: Host or student acknowledges/dismisses the lifeline
        self.client.force_authenticate(user=self.admin_user) # Host dismisses
        ack_url = f"/api/quizzes/{self.quiz.id}/hotseat-lifeline-acknowledge/"
        response = self.client.post(ack_url, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Verify attempt state is cleared and quiz timer is unpaused
        attempt = HotseatAttempt.objects.get(id=self.attempt.id)
        self.assertFalse(attempt.timer_is_paused)
        self.assertEqual(attempt.lifeline_request_status, 'none')
        self.assertEqual(attempt.pending_lifeline_type, '')


class QuizRoundConfigurationTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            email="admin_rounds@test.com",
            full_name="Admin User",
            password="adminpassword"
        )
        self.client = APIClient()

    def test_quiz_round_defaults(self):
        # Verify default value is True for all round toggles
        quiz = Quiz.objects.create(
            title="Default Rounds Quiz",
            created_by=self.admin_user
        )
        self.assertTrue(quiz.has_prelim_round)
        self.assertTrue(quiz.has_buzzer_round)
        self.assertTrue(quiz.has_fff_round)
        self.assertTrue(quiz.has_hotseat_round)

    def test_quiz_round_serializer_read_write(self):
        # Authenticate admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Create a quiz with customized round toggles via API
        url = "/api/quizzes/admin/"
        payload = {
            "title": "Custom Rounds Quiz",
            "has_prelim_round": False,
            "has_buzzer_round": True,
            "has_fff_round": False,
            "has_hotseat_round": True,
            "status": "draft"
        }
        response = self.client.post(url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertFalse(response.data["has_prelim_round"])
        self.assertTrue(response.data["has_buzzer_round"])
        self.assertFalse(response.data["has_fff_round"])
        self.assertTrue(response.data["has_hotseat_round"])

        # Fetch the created quiz to confirm DB persistence
        quiz_id = response.data["id"]
        quiz = Quiz.objects.get(id=quiz_id)
        self.assertFalse(quiz.has_prelim_round)
        self.assertTrue(quiz.has_buzzer_round)
        self.assertFalse(quiz.has_fff_round)
        self.assertTrue(quiz.has_hotseat_round)

    def test_upload_questions_for_disabled_round(self):
        # Authenticate admin
        self.client.force_authenticate(user=self.admin_user)
        
        # Create a quiz with prelim and fff disabled
        quiz = Quiz.objects.create(
            title="Buzzer/Hotseat Only Quiz",
            has_prelim_round=False,
            has_fff_round=False,
            has_buzzer_round=True,
            has_hotseat_round=True,
            created_by=self.admin_user
        )
        
        # Build mock Excel with prelim ('regular') question type
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        headers = [
            'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
            'Correct Option (A/B/C/D)', 'Marks', 
            'Question Type (regular)', 'Category', 'Trivia'
        ]
        ws.append(headers)
        ws.append([
            "What is 2+2?", "3", "4", "5", "6", "B", 1, "regular", "Math", "Addition"
        ])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "questions.xlsx"
        
        url = f"/api/quizzes/admin/{quiz.id}/upload_questions/"
        response = self.client.post(url, {"file": excel_file}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["error_count"], 1)
        self.assertEqual(response.data["success_count"], 0)
        self.assertEqual(response.data["errors"][0]["reason"], "Preliminary MCQ round is disabled for this quiz.")


class QuizSecurityLoopholesTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.school_a = School.objects.create(school_name="School A", school_code="SCHA")
        self.school_b = School.objects.create(school_name="School B", school_code="SCHB")
        
        self.program = Program.objects.create(school=self.school_a, program_name="BTech", program_code="BTECH")
        self.branch = Branch.objects.create(program=self.program, branch_name="CSE", branch_code="CSE")

        # Admin A (School A)
        self.admin_a = User.objects.create_user(
            email="admin_a@test.com",
            full_name="Admin A",
            password="adminpassword",
            college_id="ADMIN_A",
            role=User.Role.ADMIN,
            school=self.school_a,
            is_staff=True
        )
        # Admin B (School B)
        self.admin_b = User.objects.create_user(
            email="admin_b@test.com",
            full_name="Admin B",
            password="adminpassword",
            college_id="ADMIN_B",
            role=User.Role.ADMIN,
            school=self.school_b,
            is_staff=True
        )

        # Student A (School A)
        self.student_a = User.objects.create_user(
            email="student_a@test.com",
            full_name="Student A",
            password="studentpassword",
            college_id="STUD_A",
            role=User.Role.STUDENT
        )
        StudentProfile.objects.create(
            user=self.student_a,
            school=self.school_a,
            program=self.program,
            branch=self.branch,
            year="1"
        )

        # Student B (School B)
        self.student_b = User.objects.create_user(
            email="student_b@test.com",
            full_name="Student B",
            password="studentpassword",
            college_id="STUD_B",
            role=User.Role.STUDENT
        )
        StudentProfile.objects.create(
            user=self.student_b,
            school=self.school_b,
            program=self.program,
            branch=self.branch,
            year="1"
        )

        # Quiz owned by Admin A (School A), allowed schools: School A
        self.quiz = Quiz.objects.create(
            title="Quiz A",
            created_by=self.admin_a,
            host=self.admin_a,
            visible_to_students=True
        )
        self.quiz.allowed_schools.add(self.school_a)

    def test_student_team_join_duplication_restricted(self):
        # Register both students for the quiz (Student A registered for Quiz A, Student B registered for Quiz A - wait, Student B is School B, let's temporarily allow School B or register them directly)
        self.quiz.allowed_schools.add(self.school_b)
        
        reg_a = QuizRegistration.objects.create(
            student=self.student_a,
            quiz=self.quiz,
            payment_status=QuizRegistration.PaymentStatus.PAID,
            sequence_number=1,
            player_id="PLAYER 001"
        )
        reg_b = QuizRegistration.objects.create(
            student=self.student_b,
            quiz=self.quiz,
            payment_status=QuizRegistration.PaymentStatus.PAID,
            sequence_number=2,
            player_id="PLAYER 002"
        )

        from quizzes.models import Team
        # Create Team 1 with Student A as leader
        team_1 = Team.objects.create(
            name="Team 1",
            quiz=self.quiz,
            leader=self.student_a,
            member1_name=self.student_a.full_name,
            member1_email=self.student_a.email
        )

        # Student B joins Team 1 -> Should succeed
        self.client.force_authenticate(user=self.student_b)
        url = f"/api/quizzes/teams/{team_1.id}/join/"
        response = self.client.post(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Create Team 2 with Student B as leader -> Should fail because Student B is already in Team 1
        url_create = "/api/quizzes/teams/"
        payload = {
            "name": "Team 2",
            "quiz": self.quiz.id,
            "member1_email": self.student_b.email,
            "member2_email": "extra@test.com",
            "member3_email": "extra2@test.com",
            "member4_email": "extra3@test.com",
        }
        response = self.client.post(url_create, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        # Try to join another team as Student A -> Should fail because Student A is leader of Team 1
        team_3 = Team.objects.create(
            name="Team 3",
            quiz=self.quiz,
            leader=self.student_b,
            member1_name=self.student_b.full_name,
            member1_email=self.student_b.email
        )
        self.client.force_authenticate(user=self.student_a)
        url_join_3 = f"/api/quizzes/teams/{team_3.id}/join/"
        response = self.client.post(url_join_3)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_admin_cross_school_idor_restricted(self):
        # Admin B tries to access/manage Admin A's quiz
        self.client.force_authenticate(user=self.admin_b)

        # 1. Test admin buzzer init endpoint -> Should get 403 since they don't own it
        url_buzzer = f"/api/quizzes/admin/{self.quiz.id}/buzzer_init/"
        response = self.client.post(url_buzzer)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        # 2. Test edit/delete quiz via viewset when school is NOT allowed -> Should get 404 (hidden)
        url_edit = f"/api/quizzes/admin/{self.quiz.id}/"
        payload = {"title": "Hacked Title"}
        response = self.client.patch(url_edit, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

        response = self.client.delete(url_edit)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

        # 3. Test edit/delete quiz via viewset when school IS allowed -> Should get 403 Forbidden (restricted write)
        self.quiz.allowed_schools.add(self.school_b)
        response = self.client.patch(url_edit, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        response = self.client.delete(url_edit)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_student_school_boundary_quiz_detail_restricted(self):
        # Quiz is allowed for School A. Student B (School B) tries to view details
        self.client.force_authenticate(user=self.student_b)
        url = f"/api/quizzes/{self.quiz.id}/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

        # Student A (School A) should succeed
        self.client.force_authenticate(user=self.student_a)
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)






